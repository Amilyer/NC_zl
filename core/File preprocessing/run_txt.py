from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side
)
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import re
import os
import time
import chardet
import json

class FillMessage:
    """填充加工程序单数据的类（支持A-F多面工作表，初始默认A面）"""
    def __init__(self, excel_path=None, wb=None, ws=None):
        self.excel_path = excel_path
        self.wb = wb
        self.ws = ws
        self.txt_content = None
        self.txt_list = []
        self.process_times = []
        self.json_data = None
        # 统一边框样式
        self.base_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )
        self.black_border = Border(
            left=Side(style="medium", color="000000"),
            right=Side(style="medium", color="000000"),
            top=Side(style="medium", color="000000"),
            bottom=Side(style="medium", color="000000")
        )
        
    def _load_txt(self, txt_path):
        """读取txt内容并缓存"""
        if not self.txt_content:
            with open(txt_path, "rb") as f:
                raw = f.read()
                encoding = chardet.detect(raw)["encoding"]
            try:
                with open(txt_path, "r", encoding=encoding, errors="ignore") as f:
                    self.txt_content = f.read()
            except Exception:
                with open(txt_path, "r", encoding="latin1") as f:
                    self.txt_content = f.read()
        return self.txt_content
    
    def _load_json(self, json_path):
        """读取json文件内容并缓存"""
        if not self.json_data and os.path.exists(json_path):
            try:
                with open(json_path, "r", encoding="utf-8") as f:
                    self.json_data = json.load(f)
                print(f"成功加载JSON文件：{json_path}")
            except Exception as e:
                print(f"加载JSON文件失败：{e}")
                self.json_data = None
        return self.json_data
    
    def _extract_json_operation_data(self, json_path):
        """从JSON提取每道工序的Toolpath Time和Spindle RPM（增强匹配）"""
        json_data = self._load_json(json_path)
        if not json_data or "operations" not in json_data:
            print("JSON数据中无operations字段")
            return []
        
        operation_list = []
        for op in json_data["operations"]:
            op_data = {
                "toolpath_time": 0.0,
                "spindle_rpm": 0.0
            }
            for param in op.get("parameters", []):
                # 兼容大小写/空格/别名
                display_name = param.get("display_name", "").strip().lower()
                value = param.get("value", 0.0)
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    value = 0.0
                
                # 匹配Toolpath Time（兼容不同写法）
                if "toolpath time" in display_name or "加工时间" in display_name:
                    op_data["toolpath_time"] = value
                # 匹配Spindle RPM（兼容不同写法）
                elif "spindle rpm" in display_name or "主轴转速" in display_name or "转速" in display_name:
                    op_data["spindle_rpm"] = value
            operation_list.append(op_data)
        
        print(f"从JSON提取到{len(operation_list)}道工序的时间和转速数据")
        # 打印提取结果（调试用）
        for idx, op in enumerate(operation_list):
            print(f"工序{idx+1}：时间={op['toolpath_time']}分，转速={op['spindle_rpm']}rpm")
        return operation_list
    
    def _split_txt_by_multiple_blocks(self, txt_path, split_keyword="信息列表创建者 Admin"):
        """按重复出现的关键词行分块"""
        self._load_txt(txt_path)
        lines = [line.strip() for line in self.txt_content.splitlines() if line.strip()]
        
        blocks = []
        current_block = []
        split_count = 0
        
        for line in lines:
            if line == split_keyword:
                split_count += 1
                if current_block:
                    blocks.append(current_block)
                    current_block = []
                current_block.append(line)
            else:
                current_block.append(line)
        
        if current_block:
            blocks.append(current_block)
        
        print(f"共找到{split_count}个分隔行，分成{len(blocks)}个工序块")
        return blocks
    
    def _split_blocks_by_ABCDEF(self, blocks):
        """拆分工序块到A-F面：
        1. 初始默认归属A面
        2. 顺序组含A-F字母则切换对应面，后续块继承当前面
        3. 未触发的面（如DEF）则对应块列表为空
        """
        # 初始化各面的块容器
        face_blocks = {
            "A面": [], "B面": [], "C面": [], 
            "D面": [], "E面": [], "F面": []
        }
        current_face = "A面"  # 初始默认归属A面
        
        for block in blocks:
            order_group = ""
            # 提取当前块的顺序组
            for line in block:
                match = re.match(r'^(顺序组)\s*[:：\s]\s*(.+)$', line.strip())
                if match:
                    order_group = match.group(2).strip()
                    break
            
            # 检查顺序组中的A-F标识，触发归属切换
            face_matched = None
            for face_char in ["A", "B", "C", "D", "E", "F"]:
                if face_char in order_group.upper():
                    face_matched = f"{face_char}面"
                    break
            
            # 切换归属面（仅当识别到A-F字母时）
            if face_matched:
                current_face = face_matched
            
            # 将当前块添加到当前归属面的列表中
            face_blocks[current_face].append(block)
        
        # 打印拆分结果
        for face, blocks_list in face_blocks.items():
            print(f"{face}：{len(blocks_list)}个工序块")
        
        return face_blocks
    
    def _preprocess_value(self, value):
        """处理等号开头的问题"""
        if value and isinstance(value, str) and value.startswith("="):
            return f"'{value}"
        return value
    
    def _format_decimal(self, value):
        """格式化数值为两位小数（保留负号）"""
        if not value:
            return 0.00
        num_match = re.search(r"(-?\d+\.?\d*)", str(value))
        if num_match:
            num = float(num_match.group(1))
            return round(num, 2)
        return 0.00
    
    def _time_to_minutes(self, time_str):
        """将时间字符串转换为分钟数（舍去秒）"""
        if not time_str:
            return 0
        try:
            h, m, s = map(int, time_str.split(":"))
            return h * 60 + m
        except:
            return 0
    
    def _fill_system_date(self):
        """填充系统日期"""
        date_cell = self._find_excel_cell("日期")
        if date_cell:
            current_date = datetime.now().strftime("%Y-%m-%d")
            target_cell = self.ws.cell(row=date_cell.row, column=date_cell.column + 1, value=current_date)
            target_cell.border = self.base_border
            print(f"已填充{self.ws.title}系统日期：{current_date}")
    
    def _extract_valid_size(self, text):
            """辅助方法：判断是否为120*120*120形式的尺寸，是则提取标准化的L/W/T格式，否则返回None"""
            # 匹配支持的尺寸形式：
            # 120*120*120、L120*W120*T120、120×120×120、L120×W120×H120等（支持小数、空格）
            pattern = r"""
                ^\s*(\d+\.?\d*)\s*\*\s*(\d+\.?\d*)\s*\*\s*(\d+\.?\d*)\s*$
                """
            # 编译正则（忽略空格和注释，增强可读性）
            size_pattern = re.compile(pattern, re.VERBOSE)
            
            # 统一替换全角×为半角*，方便匹配
            text = text.replace("×", "*")
            match = size_pattern.search(text)
            
            if match:
                # 提取三组数字并格式化为L/W/T格式
                l, w, t = match.groups()
                return f"L{l} W{w} T{t}"
            return None

    def _fill_workpiece_size(self, txt_path):
        """从txt第一行提取工件尺寸（仅提取120*120*120形式的尺寸）"""
        self._load_txt(txt_path)
        first_line = self.txt_content.splitlines()[0].strip() if self.txt_content else ""
        
        # 调用辅助方法判断并提取有效尺寸
        workpiece_size = self._extract_valid_size(first_line)
        
        # 定位工件尺寸单元格
        size_cell = self._find_excel_cell("工件尺寸")
        if not size_cell:
            return
        
        # 根据提取结果处理单元格
        if workpiece_size:
            # 提取到有效尺寸，填充尺寸值
            target_cell = self.ws.cell(row=size_cell.row, column=size_cell.column + 1, value=workpiece_size)
            target_cell.border = self.base_border
            print(f"已填充{self.ws.title}工件尺寸：{workpiece_size}")
        else:
            # 未提取到有效尺寸，填充“未提取到尺寸”
            target_cell = self.ws.cell(row=size_cell.row, column=size_cell.column + 1, value="未提取到尺寸")
            target_cell.border = self.base_border
            print(f"{self.ws.title}未提取到有效工件尺寸（非120*120*120形式），已填充提示文本")
    
    def set_customer_dropdown(self, dropdown_configs):
        """批量设置下拉框"""
        for config in dropdown_configs:
            target_cell = config["target_cell"]
            options = config["options"]
            option_col = config.get("option_col")
            comment = config.get("comment", "")
            field_name = config.get("field_name")
            
            formula1 = ""
            if option_col:
                field_cell = f"{option_col}1"
                self.ws[field_cell] = field_name
                self.ws[field_cell].font = Font(name="微软雅黑", size=10, bold=True)
                self.ws[field_cell].border = self.base_border
                
                for idx, option in enumerate(options):
                    cell_coord = f"{option_col}{idx+2}"
                    self.ws[cell_coord] = option
                    self.ws[cell_coord].font = Font(name="微软雅黑", size=10)
                    self.ws[cell_coord].border = self.base_border
                self.ws.column_dimensions[option_col].width = 10
                formula1 = f"{option_col}2:{option_col}{max(len(options)+1, 201)}"
            else:
                formula1 = '"{}"'.format(",".join(options))
            
            dv = DataValidation(
                type="list",
                formula1=formula1,
                allow_blank=True
            )
            dv.add(self.ws[target_cell])
            self.ws.add_data_validation(dv)
            
            if len(options) > 1 and not self.ws[target_cell].value:
                self.ws[target_cell] = options[1]
                print(f"已默认填充{self.ws.title}-{target_cell}为：{options[1]}")
            
            if comment:
                from openpyxl.comments import Comment
                self.ws[target_cell].comment = Comment(comment, "提示")
            
            self.ws[target_cell].border = self.base_border
            print(f"{self.ws.title}下拉框已设置：{target_cell} → {options}")
    
    def add_process_type_remarks_dropdown(self, header_row=14):
        """程序名称非空时，给对应行的加工类型和备注加下拉框"""
        header_cols = {}
        for col in range(1, self.ws.max_column + 1):
            header_value = self.ws.cell(row=header_row, column=col).value
            if header_value in ["程序名称", "加工类型", "备注"]:
                header_cols[header_value] = col
        
        if not all(key in header_cols for key in ["程序名称", "加工类型", "备注"]):
            print(f"{self.ws.title}警告：未找到程序名称/加工类型/备注表头")
            return
        
        # 加工类型选项（AE列）
        process_type_options = ["加工类型","粗加工", "精加工", "半精加工", "钻孔", "攻丝", "铣削"]
        option_col_type = "AE"
        for idx, option in enumerate(process_type_options):
            cell_coord = f"{option_col_type}{idx+1}"
            self.ws[cell_coord] = option
            self.ws[cell_coord].font = Font(name="微软雅黑", size=10)
            self.ws[cell_coord].border = self.base_border
        self.ws.column_dimensions[option_col_type].width = 10
        formula1_type = f"{option_col_type}1:{option_col_type}{max(len(process_type_options), 20)}"

        # 备注选项（AF列）
        remarks_options = ["备注","正常加工", "需留余量", "需抛光", "需检测", "加急"]
        option_col_remarks = "AF"
        for idx, option in enumerate(remarks_options):
            cell_coord = f"{option_col_remarks}{idx+1}"
            self.ws[cell_coord] = option
            self.ws[cell_coord].font = Font(name="微软雅黑", size=10)
            self.ws[cell_coord].border = self.base_border
        self.ws.column_dimensions[option_col_remarks].width = 10
        formula1_remarks = f"{option_col_remarks}1:{option_col_remarks}{max(len(remarks_options), 20)}"

        # 遍历数据行
        for row in range(header_row + 1, self.ws.max_row + 1):
            program_name_cell = self.ws.cell(row=row, column=header_cols["程序名称"])
            if program_name_cell.value and program_name_cell.value.strip():
                # 加工类型下拉框
                process_type_cell = self.ws.cell(row=row, column=header_cols["加工类型"])
                dv_type = DataValidation(type="list", formula1=formula1_type, allow_blank=True)
                dv_type.add(process_type_cell)
                self.ws.add_data_validation(dv_type)
                if not process_type_cell.value:
                    process_type_cell.value = process_type_options[1]
                process_type_cell.border = self.base_border
                
                # 备注下拉框
                remarks_cell = self.ws.cell(row=row, column=header_cols["备注"])
                dv_remarks = DataValidation(type="list", formula1=formula1_remarks, allow_blank=True)
                dv_remarks.add(remarks_cell)
                self.ws.add_data_validation(dv_remarks)
                if not remarks_cell.value:
                    remarks_cell.value = remarks_options[1]
                remarks_cell.border = self.base_border
                
                print(f"已给{self.ws.title}第{row}行添加加工类型/备注下拉框")

    def fill_blocks_to_table(self, blocks, json_operation_data=None, header_row=14):
        """填充工序块到当前工作表"""
        # 定位表头列位置
        header_cols = {}
        for col in range(1, self.ws.max_column + 1):
            header_value = self.ws.cell(row=header_row, column=col).value
            if header_value in ["序号", "程序名称", "刀具", "刀号", "转速", "时间", 
                            "进给", "最深Z值", "加工类型", "装刀长", "余量(侧/底)", "备注", "刀柄(说明)"]:
                header_cols[header_value] = col
        
        if not header_cols:
            raise ValueError(f"{self.ws.title}第{header_row}行未找到表头")
        
        # 填充数据（仅填充实际工序，不预填充空行）
        start_data_row = header_row + 1
        total_processes = len(blocks)
        self.process_times = []
        
        # 字段正则匹配规则
        field_regex = {
            "程序名称": r'^(程序名称|工序名称)\s*[:：\s]\s*(.+)$',
            "刀具": r'^(刀具|刀具名|刀具信息)\s*[:：\s]\s*-?(.*)$',
            "刀号": r'^(刀号|刀具号)\s*[:：\s]\s*(.+)$',
            "转速": r'^(转速)\s*[:：]\s*(\d+(?:\.\d+)?)',
            "时间": r'^(时间|机床总时间|加工时间)\s*[:：\s]\s*(.+)$',
            "进给": r'^(切削|进给)\s*[:：\s]\s*(.+)$',
            "最深Z值": r'^(OPER_MIN_Z|最深Z值)\s*[:：\s]\s*(.+)$',
            "部件余量": r'^(部件余量)\s*[:：\s]\s*(.+)$',
            "底面余量": r'^(底面余量)\s*[:：\s]\s*(.+)$',
            "刀柄(说明)": r'^(刀柄|刀柄说明)\s*[:：\s]\s*(.*?)[\s:：]*$',
            "加工类型": r'^(加工类型)\s*[:：]\s*([^\s:：]+)',
        }
        
        field_mapping = {
            "程序名称": ["程序名称", "顺序组","工序名称"],
            "刀具": ["刀具", "刀具名","刀具信息"],
            "刀号": ["刀号", "刀具号" ],
            "转速": ["转速"],
            "时间": ["机床总时间", "时间", "加工时间"],
            "进给": ["切削", "进给"],
            "最深Z值": ["OPER_MIN_Z", "最深Z值"],
            "余量(侧/底)": ["部件余量", "底面余量"],
            "刀柄(说明)": ["刀柄", "刀柄说明"],
            "加工类型": ["加工类型", "工序类型"],
            "备注": ["备注", "说明"],
        }
        
        # 先清空原有数据行（避免残留）
        for row in range(start_data_row, self.ws.max_row + 1):
            for col in header_cols.values():
                self.ws.cell(row=row, column=col, value="")
                self.ws.cell(row=row, column=col).border = self.base_border
        
        # 填充实际工序
        for block_idx, block in enumerate(blocks):
            current_row = start_data_row + block_idx
            current_seq = block_idx + 1  # 仅在实际填充时生成序号
            block_data = {}
            
            # 解析区块内的键值对
            for line in block:
                line = line.strip()
                if line:
                    for field, pattern in field_regex.items():
                        match = re.match(pattern, line)
                        if match:
                            key = match.group(1).strip()
                            value = match.group(2).strip()
                            value = re.sub(r'[^\w\u4e00-\u9fa5.:/-]', '', value)
                            block_data[key] = value
                            break
            
            # 优先使用JSON的时间和转速
            toolpath_time = 0.0
            spindle_rpm = 0.0
            if json_operation_data and block_idx < len(json_operation_data):
                toolpath_time = json_operation_data[block_idx].get("toolpath_time", 0.0)
                spindle_rpm = json_operation_data[block_idx].get("spindle_rpm", 0.0)
            
            # 调试：打印当前块的JSON数据
            print(f"工序{current_seq} JSON数据：时间={toolpath_time}秒，转速={spindle_rpm}rpm")
            
            # 处理时间（JSON时间转分钟）
            if toolpath_time > 0:
                process_time = round(toolpath_time, 2)  # 分钟
                self.process_times.append(process_time)
            else:
                time_str = block_data.get("时间") or block_data.get("机床总时间")
                process_time = self._time_to_minutes(time_str)
                self.process_times.append(process_time)
            
            # 填充每个字段
            for header, col in header_cols.items():
                if header == "序号":
                    processed_value = current_seq  # 实际工序的序号
                elif header == "余量(侧/底)":
                    side_val = self._format_decimal(block_data.get("部件余量"))
                    bottom_val = self._format_decimal(block_data.get("底面余量"))
                    processed_value = f"{side_val:.2f}/{bottom_val:.2f}"
                elif header == "转速":
                    if spindle_rpm > 0:
                        processed_value = f"{spindle_rpm:.2f}"
                    else:
                        value = self._format_decimal(block_data.get("转速"))
                        processed_value = f"{value:.2f}"
                elif header == "时间":
                    if toolpath_time > 0:
                        processed_value = round(toolpath_time, 2)  # 分钟
                    else:
                        processed_value = process_time
                elif header == "刀号":
                    # 获取刀号值
                    value = None
                    for possible_key in field_mapping.get(header, [header]):
                        if possible_key in block_data:
                            value = block_data[possible_key]
                            break
                    # 空值/纯空格/值为"关"时填充0，否则保留原值
                    processed_value = "0" if (not value or not value.strip() or value.strip() == "关") else value
                else:
                    value = None
                    for possible_key in field_mapping.get(header, [header]):
                        if possible_key in block_data:
                            value = block_data[possible_key]
                            break
                    if header in ["进给", "最深Z值"]:
                        value = self._format_decimal(value)
                        processed_value = f"{value:.2f}"
                    else:
                        processed_value = self._preprocess_value(value) if value else ""

                # 写入单元格
                cell_value = processed_value if processed_value else ("0.00" if header in ["转速", "进给", "最深Z值"] else "")
                target_cell = self.ws.cell(row=current_row, column=col, value=cell_value)
                target_cell.border = self.base_border
            print(f"{self.ws.title}已填充第{current_row}行（序号{current_seq}）")
        
        # 添加黑边框
        self.add_black_border_to_valid_rows(header_row=header_row)
        print(f"\n{self.ws.title}共填充{total_processes}个工序！")
    
    def add_black_border_to_valid_rows(self, header_row=14):
        """给程序名称不为空的行添加黑边框（仅到M列）"""
        program_col = None
        for col in range(1, self.ws.max_column + 1):
            if self.ws.cell(row=header_row, column=col).value == "程序名称":
                program_col = col
                break
        
        if not program_col:
            print(f"{self.ws.title}警告：未找到“程序名称”列")
            return
        
        max_col = 13  # M列
        for row in range(header_row + 1, self.ws.max_row + 1):
            program_cell = self.ws.cell(row=row, column=program_col)
            if program_cell.value and program_cell.value.strip():
                for col in range(1, max_col + 1):
                    cell = self.ws.cell(row=row, column=col)
                    cell.border = self.black_border
                print(f"已给{self.ws.title}第{row}行添加黑边框")
    
    def calculate_total_time(self):
        """计算总时间"""
        if not self.process_times:
            print(f"{self.ws.title}无工序时间数据")
            return
        
        sum_minutes = sum(self.process_times)  # 转为分钟
        total_time = (sum_minutes * 1.2) + 20
        total_time = round(total_time, 2)
        
        total_time_cell = self._find_excel_cell("总时间(分)")
        if total_time_cell:
            target_cell = self.ws.cell(row=total_time_cell.row, column=total_time_cell.column + 1, value=total_time)
            target_cell.border = self.base_border
            print(f"{self.ws.title}已计算总时间：{total_time} 分钟")
    
    def _safe_save(self, save_path):
        """安全保存文件"""
        max_attempts = 3
        attempts = 0
        while attempts < max_attempts:
            try:
                self.wb.save(save_path)
                print(f"Excel文件已成功保存到：{save_path}")
                return
            except PermissionError:
                attempts += 1
                print(f"保存失败（文件被占用），重试第{attempts}次...")
                time.sleep(1)
        backup_path = save_path.replace(".xlsx", "_备份.xlsx")
        self.wb.save(backup_path)
        print(f"已保存到备份文件：{backup_path}")
    
    def batch_fill_downward(self, txt_path, field_configs_down):
        """批量向下填充字段"""
        self._load_txt(txt_path)
        txt_data = {}
        lines = self.txt_content.split('\n')
        for line in lines:
            line = line.strip()
            if line:
                match = re.match(r'^(\w+[\w\s]*)\s*[:：\s]\s*(.+)$', line)
                if match:
                    key = match.group(1).strip()
                    value = match.group(2).strip()
                    txt_data[key] = value
        
        for field_config in field_configs_down:
            excel_field = field_config["excel_field"]
            keyword = field_config["parse_rule"]["keyword"]
            if excel_field == "时间":
                continue
            
            value = txt_data.get(keyword, "")
            if value:
                if "余量" in excel_field or excel_field in ["转速", "进给", "最深Z值"]:
                    value = self._format_decimal(value)
                    processed_value = f"{value:.2f}"
                else:
                    processed_value = self._preprocess_value(value)
                
                excel_cell = self._find_excel_cell(excel_field)
                if excel_cell:
                    target_cell = self.ws.cell(row=excel_cell.row + 1, column=excel_cell.column)
                    if not target_cell.value:
                        target_cell.value = processed_value
                        target_cell.border = self.base_border
                        print(f"已填充{self.ws.title}「{excel_field}」下方：{processed_value}")
        print(f"\n{self.ws.title}纵向字段填充完成！")
    
    def _find_excel_cell(self, target_text):
        """查找Excel单元格"""
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value == target_text:
                    return cell
        print(f"{self.ws.title}警告：未找到「{target_text}」")
        return None
    
    def batch_fill_from_txt(self, txt_path, field_configs):
        """批量横向填充字段"""
        self._load_txt(txt_path)
        for config in field_configs:
            excel_field = config["excel_field"]
            parse_rule = config["parse_rule"]
            if excel_field in ["总时间(分)", "时间"]:
                continue
            
            value = self._parse_value(parse_rule)
            if value:
                if parse_rule["type"] == "number":
                    processed_value = f"{value:.2f}"
                else:
                    processed_value = self._preprocess_value(value)
                
                excel_cell = self._find_excel_cell(excel_field)
                if excel_cell:
                    target_cell = self.ws.cell(row=excel_cell.row, column=excel_cell.column + 1)
                    if not target_cell.value:
                        target_cell.value = processed_value
                        target_cell.border = self.base_border
                        print(f"已填充{self.ws.title}「{excel_field}」：{processed_value}")
        print(f"\n{self.ws.title}横向字段填充完成！") 
    
    def _parse_value(self, parse_rule):
        """解析txt内容"""
        txt = self.txt_content
        if parse_rule["type"] == "dimension":
            pattern = r"(\d+\.?\d*)\s*[lL].*?(\d+\.?\d*)\s*[wW].*?(\d+\.?\d*)\s*[tThH]"
            m = re.search(pattern, txt.replace("×", "x"))
            return f"L{float(m.group(1)):.2f} W{float(m.group(2)):.2f} T{float(m.group(3)):.2f}" if m else ""
        elif parse_rule["type"] == "text":
            pattern = fr"{parse_rule['keyword']}\s*[:：\s]\s*(.*?)\n"
            m = re.search(pattern, txt, re.DOTALL)
            return m.group(1).strip() if m else ""
        elif parse_rule["type"] == "number":
            pattern = fr"{parse_rule['keyword']}：(-?\d+\.?\d*)"
            m = re.search(pattern, txt)
            return round(float(m.group(1)), 2) if m else 0.00
    
    def close(self):
        """关闭Excel文件"""
        if self.wb:
            try:
                self.wb.close()
            except:
                pass
            self.wb = None

def create_process_sheet(wb, sheet_name, image_path=None):
    """在已有工作簿中创建工作表"""
    ws = wb.create_sheet(title=sheet_name)
    full_border = Border(
        left=Side(style="thin", color="000000"), 
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"), 
        bottom=Side(style="thin", color="000000")
    )
    title_font = Font(name="微软雅黑", size=16, bold=True)
    header_font = Font(name="微软雅黑", size=10, bold=True)
    normal_font = Font(name="微软雅黑", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # 标题行
    ws.merge_cells('A1:O1')
    title_cell = ws['A1']
    title_cell.value = f"加工程序单-{sheet_name}"
    title_cell.font = title_font
    title_cell.alignment = center_align
    for row in ws['A1:P1']:
        for cell in row:
            cell.border = full_border

    # 基础信息行
    ws.merge_cells('F2:M2')
    ws.merge_cells('B3:M3')
    basic_row2 = [("A2", "客户"), ("B2", ""), ("C2", "模号"), ("D2", ""), ("E2", "工件名称"), ("F2", ""), ("N2", "日期"), ("O2", "")]
    basic_row3 = [("A3", "路径"), ("B3", ""), ("N3", "工件尺寸"), ("O3", "")]
    right_info = [
        ("N4", "机床加工"), ("O4", ""), ("N5", "零件类型"), ("O5", ""), ("N6", "装夹方式"), ("O6", ""),
        ("N7", "工件X设定"), ("O7", ""), ("N8", "工件Y设定"), ("O8", ""), ("N9", "工件Z设定"), ("O9", ""),
        ("N10", "加工数量"), ("O10", ""), ("N11", "图层放置"), ("O11", ""), ("N12", "编程人员"), ("O12", ""),
        ("N13", "总时间(分)"), ("O13", "")
    ]
    ws.merge_cells('C13:M13')
    nc_path = [("A13", "NC程序文件路径"), ("B13", ""), ("C13", "")]
    
    # 写入基础信息
    all_basic_cells = basic_row2 + basic_row3 + right_info + nc_path
    for cell_coord, value in all_basic_cells:
        cell = ws[cell_coord]
        cell.value = value
        cell.font = header_font if value != "" else normal_font
        cell.alignment = center_align
        cell.border = full_border
    
    # 工序表头（第14行）
    process_header = [
        "序号", "程序名称", "刀具", "刀号", "转速", "进给",
        "最深Z值", "加工类型", "最短刀长", "余量(侧/底)", "备注", "时间", "刀柄(说明)"
    ]
    for col, header in enumerate(process_header, 1):
        cell = ws.cell(row=14, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = full_border
    
    # 初始化空行（无预填充序号）
    for row in range(15, 50):
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col, value="")
            cell.font = normal_font
            cell.alignment = left_align
            cell.border = full_border
    
    # 补充全边框
    for row in range(1, 23):
        for col in range(1, 17):
            cell = ws.cell(row=row, column=col)
            if not cell.border:
                cell.border = full_border
    
    # 调整列宽/行高
    column_widths = [5, 15, 13, 5, 8, 15, 15, 10, 8, 12, 8, 12, 8, 12, 18]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    row_heights = {1: 60,2: 30,3: 30, 4: 30, 5: 30,6: 30,7: 30,8: 30,9: 30,10: 30,11: 30,12: 30,13: 30,14: 20}
    for row, height in row_heights.items():
        ws.row_dimensions[row].height = height
    
    # 插入图片（仅当图片路径有效时）
    if image_path and os.path.exists(image_path):
        img = Image(image_path)
        img.width = 1070
        img.height = 400
        img.anchor = "A4"
        ws.add_image(img)
        print(f"{sheet_name}已插入图片：{os.path.basename(image_path)}")
    
    return ws

def get_sorted_image_paths(image_dir):
    """获取指定目录下的图片路径，按文件名排序"""
    # 支持的图片格式
    image_extensions = ('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff')
    image_paths = []
    
    if os.path.exists(image_dir):
        # 遍历目录下的文件
        for file in os.listdir(image_dir):
            # 筛选图片文件
            if file.lower().endswith(image_extensions):
                image_paths.append(os.path.join(image_dir, file))
        
        # 按文件名排序（保证顺序固定）
        image_paths.sort()
        print(f"在{image_dir}下找到{len(image_paths)}张图片，按文件名排序如下：")
        for idx, path in enumerate(image_paths):
            print(f"  第{idx+1}张：{os.path.basename(path)}")
    else:
        print(f"图片目录{image_dir}不存在")
    
    return image_paths

def extract_process_filename_from_txt(txt_path):
    """从TXT中提取加工文件名"""
    if not os.path.exists(txt_path):
        print(f"TXT文件不存在，使用默认文件名")
        return "加工程序单"
    
    try:
        # 读取TXT并自动检测编码
        with open(txt_path, "rb",) as f:
            raw = f.read()
            encoding = chardet.detect(raw)["encoding"] or "gbk"
        
        with open(txt_path, "r", encoding=encoding, errors="ignore") as f:
            # 遍历所有行，找包含"当前工作部件"的行
            for line in f:
                line = line.strip()
                if "当前工作部件" in line:
                    # 匹配路径中的文件名
                    path_match = re.search(r"([^\\/]+)\s*$", line)
                    if path_match:
                        full_filename = path_match.group(1)
                        # 去掉最后一个扩展名（.prt），保留核心名
                        filename_without_last_ext = os.path.splitext(full_filename)[0]
                        return filename_without_last_ext
        
        # 未找到则返回默认名
        return "加工程序单"
    except Exception as e:
        print(f"提取加工文件名失败：{e}，使用默认文件名")
        return "加工程序单"

def main(txt_path, json_path, image_dir, excel_save_dir):
    """主函数：核心功能实现"""

    # 验证基础路径
    if not os.path.exists(txt_path):
        print(f"警告：TXT文件不存在 - {txt_path}")
    if not os.path.exists(json_path):
        print(f"警告：JSON文件不存在 - {json_path}")
    
    # 创建Excel存储目录（不存在则自动创建）
    if not os.path.exists(excel_save_dir):
        os.makedirs(excel_save_dir)
        print(f"已创建Excel存储目录：{excel_save_dir}")
    
    try:
        # 提取加工文件名，生成Excel完整路径
        process_filename = extract_process_filename_from_txt(txt_path)
        excel_filename = f"{process_filename}.xlsx"
        excel_full_path = os.path.join(excel_save_dir, excel_filename)
        
        # 获取排序后的图片路径列表
        image_paths = get_sorted_image_paths(image_dir)
        
        # 创建工作簿和A-F面工作表
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # 定义6个面的顺序
        faces = ["A面", "B面", "C面", "D面", "E面", "F面"]
        
        # 初始化6个面的工作表和填充器
        face_sheets = {}
        face_fillers = {}
        for idx, face in enumerate(faces):
            # 为每个面匹配对应的图片路径（前6张图片）
            image_path = image_paths[idx] if idx < len(image_paths) else None
            ws = create_process_sheet(wb, face, image_path=image_path)
            face_sheets[face] = ws
            face_fillers[face] = FillMessage(wb=wb, ws=ws)
        
        # 设置下拉框（所有面统一配置）
        dropdown_configs = [
            {"target_cell": "B2", "options": ["海尔", "海信", "重普", "美的"], "option_col": "X", "field_name": "客户"},
            {"target_cell": "O4", "options": ["立式加工中心", "卧式加工中心", "五轴加工中心"], "option_col": "Y", "field_name": "机床加工"},
            {"target_cell": "O5", "options": ["模具型芯", "模具型腔", "五金零件", "塑胶零件"], "option_col": "Z", "field_name": "零件类型"},
            {"target_cell": "O6", "options": ["虎钳装夹", "压板装夹", "工装夹具"], "field_name": "装夹方式"},
            {"target_cell": "O7", "options": ["按图示", "单边", "按图示", "孔分中"], "option_col": "AA", "field_name": "工件X设定"},
            {"target_cell": "O8", "options": ["按图示", "单边", "按图示", "孔分中"], "option_col": "AB", "field_name": "工件Y设定"},
            {"target_cell": "O9", "options": ["按图示", "自定义图层1", "自定义图层2"], "option_col": "AC", "field_name": "工件Z设定"},
            {"target_cell": "O12", "options": ["张三", "李四", "王五"], "option_col": "AD", "field_name": "编程人员"}
        ]
        
        for face, filler in face_fillers.items():
            print(f"\n===== {face} 下拉框配置 =====")
            filler.set_customer_dropdown(dropdown_configs)
            # 填充系统日期（所有面都填）
            filler._fill_system_date()
        
        # 读取JSON工序数据
        json_operation_data = face_fillers["A面"]._extract_json_operation_data(json_path) if os.path.exists(json_path) else []
        
        # 处理TXT数据并填充Excel
        if os.path.exists(txt_path):
            # 定义基础信息配置
            field_configs = [
                {"excel_field": "客户", "parse_rule": {"type": "text", "keyword": "客户"}},
                {"excel_field": "模号", "parse_rule": {"type": "text", "keyword": "模号"}},
                {"excel_field": "工件名称", "parse_rule": {"type": "text", "keyword": "工件名称"}},
                {"excel_field": "路径", "parse_rule": {"type": "text", "keyword": "当前工作部件"}},
            ]
            field_configs_down = [
                {"excel_field": "程序名称", "parse_rule": {"type": "text", "keyword": "顺序组"}},
                {"excel_field": "刀具", "parse_rule": {"type": "text", "keyword": "刀具名"}},
                {"excel_field": "刀号", "parse_rule": {"type": "text", "keyword": "刀具号"}},
            ]
            
            # 所有面填充：工件尺寸 + 横向基础信息（客户/模号等）
            for face, filler in face_fillers.items():
                filler._fill_workpiece_size(txt_path)
                filler.batch_fill_from_txt(txt_path, field_configs)
            
            # 拆分原始TXT块
            blocks = face_fillers["A面"]._split_txt_by_multiple_blocks(txt_path)
            
            # 给每个原始TXT块绑定对应索引的JSON数据
            blocks_with_json = []
            for idx, block in enumerate(blocks):
                json_data = json_operation_data[idx] if idx < len(json_operation_data) else {"toolpath_time":0.0, "spindle_rpm":0.0}
                blocks_with_json.append({"block": block, "json": json_data})
            
            # 打印绑定结果（调试用）
            print("\n=== TXT块-JSON绑定结果 ===")
            for idx, item in enumerate(blocks_with_json):
                print(f"原始块{idx+1} → JSON时间：{item['json']['toolpath_time']}秒，转速：{item['json']['spindle_rpm']}rpm")
            
            # 给带JSON的TXT块分配面归属
            face_blocks_with_json = {
                "A面": [], "B面": [], "C面": [], 
                "D面": [], "E面": [], "F面": []
            }
            current_face = "A面"
            
            for item in blocks_with_json:
                block = item["block"]
                json_data = item["json"]
                
                # 提取顺序组
                order_group = ""
                for line in block:
                    match = re.match(r'^(顺序组)\s*[:：\s]\s*(.+)$', line.strip())
                    if match:
                        order_group = match.group(2).strip()
                        break
                
                # 识别面归属
                face_matched = None
                for face_char in ["A", "B", "C", "D", "E", "F"]:
                    if face_char in order_group.upper():
                        face_matched = f"{face_char}面"
                        break
                if face_matched:
                    current_face = face_matched
                
                # 添加到对应面
                face_blocks_with_json[current_face].append({"block": block, "json": json_data})
            
            # 打印面分配结果
            print("\n=== 各面绑定结果 ===")
            for face, items in face_blocks_with_json.items():
                if items:
                    print(f"{face}：")
                    for idx, item in enumerate(items):
                        print(f"  工序{idx+1} → TXT块：{item['block'][0][:20]}... | JSON时间：{item['json']['toolpath_time']}秒 | 转速：{item['json']['spindle_rpm']}rpm")
            
            # 按面填充数据
            for face, items in face_blocks_with_json.items():
                if items:
                    print(f"\n===== 填充{face}工作表 =====")
                    filler = face_fillers[face]
                    # 提取当前面的块和JSON数据
                    blocks_list = [item["block"] for item in items]
                    json_slice = [item["json"] for item in items]
                    # 填充工序
                    filler.batch_fill_downward(txt_path, field_configs_down)
                    filler.fill_blocks_to_table(blocks=blocks_list, json_operation_data=json_slice)
                    filler.calculate_total_time()
                    filler.add_process_type_remarks_dropdown()
                else:
                    print(f"\n{face}无工序块，跳过填充")
        
        # 安全保存Excel文件
        face_fillers["A面"]._safe_save(excel_full_path)
        print(f"\n===== 操作完成 ======")
        print(f"Excel文件存储路径：{excel_full_path}")
        print(f"包含工作表：{wb.sheetnames}")
    
    except Exception as e:
        print(f"\n操作失败：{str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        # 关闭所有填充器
        for filler in face_fillers.values():
            filler.close()

if __name__ == "__main__":
    # 配置项（可根据实际需求修改）
    txt_path = r"E:\get_txt\NX_Feature_Info.txt"          # NX生成的TXT路径
    json_path = r"E:\get_txt\data.json"
    
                     # JSON数据路径
    image_dir = r"E:\screen-shot"                        # 图片目录路径
    excel_save_dir = r"E:\加工程序单"                    # Excel存储目录
    
    # 执行主函数
    main(txt_path, json_path, image_dir, excel_save_dir)