import NXOpen
import NXOpen.UF
import NXOpen.CAM
import os
import threading
import time
import ctypes
import datetime
import traceback
import json
import re
import openpyxl  # 用于Excel生成

# ==================================================================================
# 统一配置项（整合所有配置，集中管理）
# ==================================================================================
class Config:
    # ---------------- 基础文件配置 ----------------
    # 部件文件路径
    PART_PATH = r"C:\Projects\NC\output\06_CAM\Final_CAM_PRT\UP-01.prt"
    # 过切检查后另存路径（仅保留此文件）
    GAUGE_CHECK_SAVE_PATH = r"C:\Users\Admin\Desktop\code2\guoqie.prt"
    # ---------------- 日志与中间文件 ----------------
    # 调试日志
    DEBUG_LOG_PATH = r"C:\Users\Admin\Desktop\code2\NX_Debug_Log.txt"
    # 过切检查输出的TXT文件 (作为Excel生成的输入)
    OUTPUT_PATH = r"C:\Users\Admin\Desktop\code2\NX_Feature_Info.txt"
    
    # ---------------- JSON参数导出配置 ----------------
    # JSON导出目录
    JSON_EXPORT_BASE_DIR = r"C:\Users\Admin\Desktop\code2"
    # 导出的文件名
    JSON_FILENAME = "data.json"
    # 完整的JSON路径 (用于Excel读取)
    JSON_FULL_PATH = os.path.join(JSON_EXPORT_BASE_DIR, JSON_FILENAME)

    # ---------------- Excel报告配置 ----------------
    # 最终生成的Excel报告路径
    EXCEL_REPORT_PATH = r"C:\Users\Admin\Desktop\code2\NX_CAM_过切检查模板.xlsx"

    # ---------------- 运行配置 ---------------- 
    # 启用日志
    ENABLE_LOG = True
    # 关闭刀轨生成后的自动保存
    AUTO_SAVE_TOOLPATH = False 

    # ---------------- 加特林线程配置 (用于处理弹窗) ----------------
    HEART_BEAT_EVERY = 10
    KEY_INTERVAL = 0.3
    BETWEEN_OPS = 0.5


# ==================================================================================
# 1. 工具函数：日志、打印、部件检查、保存
# ==================================================================================
def write_debug_log(message):
    """强力日志写入（直接写硬盘）"""
    if not Config.ENABLE_LOG:
        return
    try:
        timestamp = datetime.datetime.now().strftime("%H:%M:%S.%f")[:-3]
        with open(Config.DEBUG_LOG_PATH, "a", encoding="utf-8", buffering=1) as f:
            f.write(f"[{timestamp}] {message}\n")
            f.flush()
            os.fsync(f.fileno())
    except:
        pass

def print_to_info_window(message):
    """输出到NX信息窗口和日志文件"""
    if not Config.ENABLE_LOG:
        return
    try:
        theSession = NXOpen.Session.GetSession()
        theSession.ListingWindow.Open()
        theSession.ListingWindow.WriteLine(str(message))
        theSession.LogFile.WriteLine(str(message))
    except:
        pass # 防止无界面模式报错

def check_part_load_status(load_status):
    """检查部件加载状态"""
    try:
        if load_status == 0:
            return True
    except:
        pass
    try:
        session = NXOpen.Session.GetSession()
        if session.Parts.Work is not None:
            return True
    except:
        pass
    return False

def save_part(new_path: str = None) -> bool:
    """保存部件：指定路径则另存，否则原地保存"""
    session = NXOpen.Session.GetSession()
    work_part = session.Parts.Work
    if work_part is None:
        write_debug_log("保存失败：无有效工作部件")
        return False
    try:
        if new_path:
            work_part.SaveAs(new_path)
            write_debug_log(f"部件已另存到: {new_path}")
            print_to_info_window(f"✅ 部件已另存到 → {new_path}")
        else:
            uf_session = NXOpen.UF.UFSession.GetUFSession()
            uf_session.Part.SaveAll()
            write_debug_log("部件原地保存完成")
            print_to_info_window("✅ 部件保存完成")
        return True
    except Exception as e:
        err_msg = f"保存部件失败: {e}"
        write_debug_log(err_msg)
        print_to_info_window(f"❌ {err_msg}")
        return False

# ==================================================================================
# 2. 刀轨生成器类
# ==================================================================================
class ToolpathGeneratorMacro:
    """刀轨生成器 - 为CAM操作生成刀轨"""
    def __init__(self, session, work_part):
        self.session = session
        self.work_part = work_part
        self.success_count = 0
        self.failed_count = 0
        self.test_results = []

    def print_log(self, message, level="INFO"):
        """打印带时间戳的日志"""
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        level_symbols = {
            "INFO": "ℹ️", "SUCCESS": "✅", "ERROR": "❌",
            "WARN": "⚠️", "DEBUG": "🔍", "START": "🚀", "END": "🏁"
        }
        symbol = level_symbols.get(level, "•")
        log_msg = f"[{timestamp}] {symbol} {message}"
        print(log_msg, flush=True)
        write_debug_log(log_msg)

    def print_separator(self, char="=", length=60):
        sep = char * length
        print(sep, flush=True)
        write_debug_log(sep)

    def print_header(self, title):
        self.print_separator()
        self.print_log(f"  {title}", "START")
        self.print_separator()

    def switch_to_manufacturing(self):
        try:
            self.session.ApplicationSwitchImmediate("UG_APP_MANUFACTURING")
            self.print_log("已切换到加工环境", "SUCCESS")
        except Exception as e:
            self.print_log(f"切换加工环境警告: {e}", "WARN")

    def get_all_operations(self):
        operations = []
        try:
            for operation in self.work_part.CAMSetup.CAMOperationCollection:
                operations.append(operation)
            self.print_log(f"找到 {len(operations)} 个CAM操作", "INFO")
        except Exception as e:
            self.print_log(f"获取操作列表失败: {e}", "ERROR")
        return operations

    def generate_toolpath(self, operation):
        op_name = operation.Name
        try:
            # 设置Undo标记，虽然后面不一定会撤销，但由于API要求
            mark_id = self.session.SetUndoMark(
                NXOpen.Session.MarkVisibility.Visible,
                f"Generate Tool Path - {op_name}"
            )
            objects = [NXOpen.CAM.CAMObject.Null] * 1
            objects[0] = operation
            self.work_part.CAMSetup.GenerateToolPath(objects)
            self.print_log(f"刀轨生成完成: {op_name}", "SUCCESS")
            self.success_count += 1
            self.test_results.append({
                "name": op_name, "status": "Success", "message": "刀轨生成成功"
            })
            return True
        except Exception as e:
            err_msg = f"刀轨生成失败 - {op_name}: {e}"
            self.print_log(err_msg, "ERROR")
            write_debug_log(f"刀轨生成异常堆栈: {traceback.format_exc()}")
            self.failed_count += 1
            self.test_results.append({
                "name": op_name, "status": "Failed", "error": str(e)
            })
            return False

    def generate_all_toolpaths(self):
        self.print_header("NX CAM 刀轨生成流程")
        self.switch_to_manufacturing()
        operations = self.get_all_operations()

        if not operations:
            self.print_log("没有找到任何CAM操作", "WARN")
            return

        # 列出所有操作
        for i, op in enumerate(operations, 1):
            self.print_log(f"  {i}. {op.Name}")

        self.print_separator("-")
        self.print_log("开始生成刀轨...", "START")
        for operation in operations:
            self.generate_toolpath(operation)

        self.print_summary()

    def print_summary(self):
        total = self.success_count + self.failed_count
        success_rate = (self.success_count / total * 100) if total > 0 else 0
        self.print_separator("=")
        summary = f"""
  刀轨生成汇总
  ----------------------------------------
  总操作数:   {total}
  成功:       {self.success_count} ✅
  失败:       {self.failed_count} ❌
  成功率:     {success_rate:.1f}%
        """.strip()
        print(summary, flush=True)
        write_debug_log(summary)
        self.print_separator("=")

# ==================================================================================
# 3. 加特林线程类 (模拟按键处理弹窗)
# ==================================================================================
class EnterSpammer(threading.Thread):
    def __init__(self):
        super().__init__()
        self._stop_event = threading.Event()
        self.daemon = True

    def run(self):
        user32 = ctypes.windll.user32
        VK_RETURN = 0x0D
        write_debug_log("加特林：线程启动")
        counter = 0
        while not self._stop_event.is_set():
            user32.keybd_event(VK_RETURN, 0, 0, 0)
            user32.keybd_event(VK_RETURN, 0, 2, 0)
            counter += 1
            if counter % Config.HEART_BEAT_EVERY == 0:
                write_debug_log(f"加特林：已点击 {counter} 次")
            time.sleep(Config.KEY_INTERVAL)
        write_debug_log("加特林：线程停止")

    def stop(self):
        self._stop_event.set()

# ==================================================================================
# 4. 过切检查核心流程
# ==================================================================================
def process_gauge_check():
    """执行过切检查流程"""
    write_debug_log(">>> 开始执行过切检查流程 <<<")
    print_to_info_window("===== 开始过切检查 =====")

    session = NXOpen.Session.GetSession()
    workPart = session.Parts.Work
    if workPart is None:
        err_msg = "❌ 无有效工作部件，过切检查终止"
        print_to_info_window(err_msg)
        write_debug_log(err_msg)
        return False

    # 切换加工模块
    try:
        session.ApplicationSwitchImmediate("UG_APP_MANUFACTURING")
        uf = NXOpen.UF.UFSession.GetUFSession()
        uf.Cam.InitSession()
        print_to_info_window("✅ 加工模块初始化成功")
    except Exception as e:
        print_to_info_window(f"⚠️ 加工模块切换警告: {e}")

    # 初始化输出窗口 (这里也作为后续Excel解析的源文件)
    list_window = session.ListingWindow
    list_window.SelectDevice(NXOpen.ListingWindow.DeviceType.File, Config.OUTPUT_PATH)
    list_window.Open()
    infoTool = session.Information
    camSetup = workPart.CAMSetup

    # 获取工序列表
    try:
        operations = [op for op in camSetup.CAMOperationCollection]
    except Exception as e:
        write_debug_log(f"获取 CAMOperationCollection 失败: {e}")
        list_window.Close()
        return False

    # 启动加特林线程
    spammer = EnterSpammer()
    spammer.start()

    count_success = 0
    try:
        # 遍历工序执行过切检查
        for i, op in enumerate(operations):
            op_name = op.Name
            op_type = str(type(op))

            # 过滤钻孔类工序
            if any(x in op_type for x in ("Drill", "Hole", "Point")):
                continue

            msg = f"准备处理第 {i+1} 个工序: [{op_name}]"
            list_window.WriteLine(msg)
            write_debug_log(msg)

            try:
                write_debug_log(f"DEBUG: 正在调用 GougeCheck ({op_name}) ...")
                camSetup.GougeCheck([op], False)
                write_debug_log(f"DEBUG: GougeCheck 返回 ({op_name})")
                infoTool.DisplayCamObjectsDetails([op])

                list_window.WriteLine(f"✅ 成功: {op_name}")
                write_debug_log(f"SUCCESS: {op_name} 过切检查完毕")
                count_success += 1
            except Exception as e:
                err_msg = f"❌ 失败 {op_name}: {str(e)}"
                list_window.WriteLine(err_msg)
                write_debug_log(err_msg)

            time.sleep(Config.BETWEEN_OPS)

        # 保存过切检查后的部件（仅保留此文件）
        print_to_info_window("正在保存过切检查后的部件...")
        save_success = save_part(Config.GAUGE_CHECK_SAVE_PATH)
        if not save_success:
            print_to_info_window("⚠ 部件保存失败，请手动保存")

        return True
    finally:
        # 停止加特林线程
        spammer.stop()
        spammer.join()
        list_window.WriteLine(f"过切检查完成。成功: {count_success} 个")
        list_window.Close()
        write_debug_log(">>> 过切检查流程结束 <<<")

# ==================================================================================
# 5. NX工序参数导出器
# ==================================================================================
class NXOperationParamExporter:
    """NX工序参数导出器，用于批量扫描并汇总所有工序参数到单个JSON文件"""

    def __init__(self, session=None, work_part=None):
        self.theSession = session or NXOpen.Session.GetSession()
        self.theUFSession = NXOpen.UF.UFSession.GetUFSession()
        self.workPart = work_part or self.theSession.Parts.Work
        self.lw = self.theSession.ListingWindow
        
        self.param_dictionary = self._get_param_dictionary()
        self.summary_data = {}
        self.batch_timestamp = time.strftime("%Y%m%d_%H%M%S", time.localtime())
        self.success_count = 0
        self.fail_count = 0

    def _get_param_dictionary(self):
        return {
            116: "Suppressed",
            124: "Toolpath Time"




















            

        }

    def init_environment(self):
        self.lw.Open()
        self.lw.WriteLine("正在启动工序参数导出程序...")
        write_debug_log("正在启动工序参数导出程序...")

    def get_all_operations(self):
        camSetup = self.workPart.CAMSetup
        if camSetup is None:
            self.lw.WriteLine("\n【错误】当前部件中未检测到CAM加工环境！")
            return []

        opCollection = camSetup.CAMOperationCollection
        operations = [op for op in opCollection]
        if not operations:
            self.lw.WriteLine("\n【错误】当前部件中未检测到任何工序！")
        else:
            self.lw.WriteLine(f"\n检测到 {len(operations)} 个工序，开始批量扫描参数...")
        return operations

    def scan_operation_params(self, op):
        obj_tag = op.Tag
        collected_params = []
        # 扫描参数范围：1-10000
        for index in range(1, 10000):
            val = None
            val_type = "Unknown"
            # Try Double
            if val is None:
                try:
                    val = self.theUFSession.Param.AskDoubleValue(obj_tag, index)
                    val_type = "Double"
                    if val == 0.0 and index not in self.param_dictionary: val = None
                except: pass
            # Try Int
            if val is None:
                try:
                    val = self.theUFSession.Param.AskIntValue(obj_tag, index)
                    val_type = "Int"
                    if val == 0 and index not in self.param_dictionary: val = None
                except: pass
            # Try String
            if val is None:
                try:
                    val = self.theUFSession.Param.AskStringValue(obj_tag, index)
                    val_type = "String"
                    if val == "": val = None
                except: pass
            # Try Tag
            if val is None:
                try:
                    val = self.theUFSession.Param.AskTagValue(obj_tag, index)
                    val_type = "Tag"
                    if val == NXOpen.Tag.Null: val = None
                except: pass

            if val is not None:
                display_name = self.param_dictionary.get(index, f"UNKNOWN_ID_{index}")
                collected_params.append({
                    "id": index,
                    "display_name": display_name,
                    "type": val_type,
                    "value": val
                })
        return collected_params

    def build_summary_data(self):
        nx_version = "NX"
        try:
            nx_version = self.theSession.EnvironmentInformation.Version
        except: pass

        self.summary_data = {
            "batch_meta": {
                "export_time": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
                "nx_version": nx_version,
                "part_name": self.workPart.Name if self.workPart else "未知部件",
                "total_operations": 0,
                "batch_timestamp": self.batch_timestamp,
                "success_operations": 0,
                "fail_operations": 0
            },
            "operations": []
        }

    def process_operations(self, operations):
        self.summary_data["batch_meta"]["total_operations"] = len(operations)
        for idx, op in enumerate(operations, 1):
            op_name = op.Name
            self.lw.WriteLine(f"\n[{idx}/{len(operations)}] 正在处理工序：{op_name}")
            try:
                collected_params = self.scan_operation_params(op)
                self.summary_data["operations"].append({
                    "operation_name": op_name,
                    "operation_type": type(op).__name__,
                    "total_params": len(collected_params),
                    "parameters": collected_params,
                    "status": "success"
                })
                self.success_count += 1
                self.summary_data["batch_meta"]["success_operations"] = self.success_count
            except Exception as e:
                self.summary_data["operations"].append({
                    "operation_name": op_name,
                    "status": "failed",
                    "error_message": str(e)
                })
                self.fail_count += 1
                self.summary_data["batch_meta"]["fail_operations"] = self.fail_count
                self.lw.WriteLine(f"    【处理失败】: {str(e)}")

    def save_summary_file(self):
        base_dir = Config.JSON_EXPORT_BASE_DIR
        if not os.path.exists(base_dir):
            try:
                os.makedirs(base_dir)
            except OSError as e:
                self.lw.WriteLine(f"\n【错误】无法创建目录 {base_dir}：{str(e)}")
                return False

        full_path = Config.JSON_FULL_PATH
        try:
            with open(full_path, "w", encoding='utf-8') as f:
                json.dump(self.summary_data, f, indent=4, ensure_ascii=False)
            self.lw.WriteLine(f"\n成功保存JSON文件: {full_path}")
            write_debug_log(f"JSON参数文件已保存到：{full_path}")
            return True
        except Exception as e:
            self.lw.WriteLine(f"\n【保存失败】: {str(e)}")
            return False

    def run(self):
        self.init_environment()
        operations = self.get_all_operations()
        if not operations: return
        self.build_summary_data()
        self.process_operations(operations)
        self.save_summary_file()

# ==================================================================================
# 6. Excel 报告生成器 (数据处理逻辑)
# ==================================================================================
class ExcelReportGenerator:
    """处理文本和JSON数据，生成Excel报告"""
    
    @staticmethod
    def split_txt_by_generated_on(file_path):
        blocks = []
        current_block = []
        pattern = re.compile(r'\s*GENERATED\s+ON\s*', re.IGNORECASE)

        if not os.path.exists(file_path):
            print(f"ExcelGenerator: TXT文件不存在 {file_path}")
            return []

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line_stripped = line.rstrip('\n')
                    if pattern.search(line_stripped):
                        if current_block:
                            blocks.append(current_block)
                        current_block = [line_stripped]
                    else:
                        current_block.append(line_stripped)
                if current_block:
                    blocks.append(current_block)
            return blocks
        except Exception as e:
            print(f"ExcelGenerator: 读取文件错误 {e}")
            return []

    @staticmethod
    def extract_process_info(block):
        # 1. 提取工序名
        process_pattern = re.compile(r'(Operation Name)\s*[:：\s]\s*(.+)', re.IGNORECASE)
        # 2. 提取过切状态
        gouge_pattern = re.compile(r'(Gouge Check Status)\s*[:：\s]\s*([^；，。\n]+)', re.IGNORECASE)

        process_name = "未知工序"
        gouge_status = "None"

        for line in block:
            if process_name == "未知工序":
                process_match = process_pattern.search(line)
                if process_match:
                    process_name = process_match.group(2).strip()
            if gouge_status == "None":
                gouge_match = gouge_pattern.search(line)
                if gouge_match:
                    gouge_status = gouge_match.group(2).strip()
            if process_name != "未知工序" and gouge_status != "None":
                break
        return process_name, gouge_status

    @staticmethod
    def extract_toolpath_time_from_json(json_path):
        toolpath_times = []
        if not os.path.exists(json_path):
            print(f"ExcelGenerator: JSON文件不存在 {json_path}")
            return toolpath_times
        
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            operations = data.get('operations', [])
            for op in operations:
                toolpath_time = 0.0
                parameters = op.get('parameters', [])
                for param in parameters:
                    display_name = param.get('display_name', '').strip().lower()
                    if 'toolpath time' in display_name or '加工时间' in display_name:
                        try:
                            toolpath_time = float(param.get('value', 0.0))
                        except:
                            toolpath_time = 0.0
                        break
                toolpath_times.append(toolpath_time)
            print(f"从JSON中提取到 {len(toolpath_times)} 个Toolpath Time值")
            return toolpath_times
        except Exception as e:
            print(f"ExcelGenerator: JSON解析错误 {e}")
            return toolpath_times

    @staticmethod
    def write_to_excel(excel_path, process_names, gouge_statuses, toolpath_statuses, part_file_path):
        from openpyxl.utils import get_column_letter  # 局部导入避免影响顶部
        try:
            if os.path.exists(excel_path):
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "CAM过切检查结果"
                ws.cell(row=1, column=1, value="文件路径")    # A1: 新增
                ws.cell(row=1, column=2, value="工序名")
                ws.cell(row=1, column=3, value="是否过切")
                ws.cell(row=1, column=4, value="是否生成刀路")
                print(f"已创建新Excel文件：{excel_path}")

            # 确保表头存在
            if ws.cell(row=1, column=1).value != "文件路径":
                ws.cell(row=1, column=1, value="文件路径")

            process_col = 2
            gouge_col = 3
            toolpath_col = 4
            start_row = 2

            # 写入数据
            max_len = max(len(process_names), len(gouge_statuses), len(toolpath_statuses))
            for idx in range(max_len):
                # 写入文件路径
                ws.cell(row=start_row + idx, column=1, value=part_file_path)
                
                if idx < len(process_names):
                    ws.cell(row=start_row + idx, column=process_col, value=process_names[idx])
                if idx < len(gouge_statuses):
                    ws.cell(row=start_row + idx, column=gouge_col, value=gouge_statuses[idx])
                if idx < len(toolpath_statuses):
                    ws.cell(row=start_row + idx, column=toolpath_col, value=toolpath_statuses[idx])
                else:
                    ws.cell(row=start_row + idx, column=toolpath_col, value="未知")

            # 自动调整列宽
            for column_cells in ws.columns:
                length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
                # 适当增加一点宽度作为缓冲，并设置上限防止过宽
                adjusted_width = min((length + 2) * 1.2, 50) 
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

            wb.save(excel_path)
            print(f"Excel报告生成成功！共写入 {max_len} 条数据。")
            print(f"路径: {excel_path}")
        except PermissionError:
            print("错误：Excel文件已被占用，请关闭后重试。")
        except Exception as e:
            print(f"Excel写入失败：{e}")
            import traceback
            traceback.print_exc()

    @classmethod
    def run_report_generation(cls):
        print("\n===== 开始生成Excel报告 =====")
        write_debug_log(">>> 开始生成Excel报告 <<<")
        
        # 1. 处理TXT
        txt_blocks = cls.split_txt_by_generated_on(Config.OUTPUT_PATH)
        if not txt_blocks:
            print("提示: 未从TXT中提取到有效块，跳过")
            return

        process_info = [cls.extract_process_info(block) for block in txt_blocks]
        process_names = [info[0] for info in process_info]
        gouge_statuses = [info[1] for info in process_info]

        # 2. 处理JSON
        toolpath_times = cls.extract_toolpath_time_from_json(Config.JSON_FULL_PATH)

        # 3. 对齐数据 (以JSON为准)
        min_len = len(toolpath_times)
        if len(process_names) > min_len:
            print(f"ExcelGenerator: TXT数据({len(process_names)}) 多于 JSON数据({min_len})，正在截断...")
            process_names = process_names[:min_len]
            gouge_statuses = gouge_statuses[:min_len]
        elif len(toolpath_times) > len(process_names):
            print(f"ExcelGenerator: JSON数据({len(toolpath_times)}) 多于 TXT数据({len(process_names)})，部分行将为空")

        # 4. 生成刀路生成状态
        toolpath_statuses = ["是" if t > 0 else "否" for t in toolpath_times]

        # 5. 写入Excel
        cls.write_to_excel(Config.EXCEL_REPORT_PATH, process_names, gouge_statuses, toolpath_statuses, Config.PART_PATH)
        write_debug_log(">>> Excel报告生成完毕 <<<")


# ==================================================================================
# 7. 主工作流
# ==================================================================================
def main():
    """整合主流程：刀轨 -> 过切 -> JSON -> Excel"""
    # 清空旧日志
    if os.path.exists(Config.DEBUG_LOG_PATH):
        try: os.remove(Config.DEBUG_LOG_PATH)
        except: pass
    
    write_debug_log(">>> 整合流程开始运行 (Python Integrated) <<<")
    session = NXOpen.Session.GetSession()

    # 1. 检查部件
    if not os.path.exists(Config.PART_PATH):
        print_to_info_window(f"❌ 错误: 部件不存在 {Config.PART_PATH}")
        return False

    # 2. 打开部件
    print_to_info_window(f"正在打开部件: {Config.PART_PATH}")
    base_part, load_status = session.Parts.OpenBaseDisplay(Config.PART_PATH)
    if not check_part_load_status(load_status):
        print_to_info_window("❌ 打开部件失败")
        return False
    work_part = session.Parts.Work

    # 3. 生成刀轨
    generator = ToolpathGeneratorMacro(session, work_part)
    generator.generate_all_toolpaths()
    generator.print_log("刀轨生成结束", "END")

    # 4. 过切检查 (生成 TXT)
    process_gauge_check()

    # 5. 导出JSON (生成 JSON)
    print_to_info_window("\n===== 开始导出工序参数为JSON =====")
    exporter = NXOperationParamExporter(session, work_part)
    exporter.run()

    # 6. 生成Excel报告 (读取 TXT + JSON -> Excel)
    # 注意：这里使用普通print，因为这部分逻辑通常在控制台可见即可，或者也可以封装进print_to_info_window
    ExcelReportGenerator.run_report_generation()

    print_to_info_window("\n===== 整合流程全部完成 =====")
    write_debug_log(">>> 整合流程全部结束 <<<")
    return True

if __name__ == "__main__":
    main()

