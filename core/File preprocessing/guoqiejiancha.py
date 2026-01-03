import NXOpen
import NXOpen.UF
import NXOpen.CAM
import os
import threading
import time
import ctypes
import datetime
import traceback
import json
import re
import openpyxl  # ç”¨äºExcelç”Ÿæˆ

# ==================================================================================
# 1. å·¥å…·å‡½æ•°ï¼šéƒ¨ä»¶æ£€æŸ¥ã€ä¿å­˜
# ==================================================================================
def check_part_load_status(load_status):











    """æ£€æŸ¥éƒ¨ä»¶åŠ è½½çŠ¶æ€"""
    try:
        if load_status == 0:
            return True
    except:
        pass
    try:
        session = NXOpen.Session.GetSession()
        if session.Parts.Work is not None:
            return True
    except:
        pass
    return False

def save_part(new_path: str) -> bool:
    """ä¿å­˜éƒ¨ä»¶ï¼šæŒ‡å®šè·¯å¾„åˆ™å¦å­˜"""
    session = NXOpen.Session.GetSession()
    work_part = session.Parts.Work
    if work_part is None:
        return False
    try:
        # ç¡®ä¿ä¿å­˜è·¯å¾„çš„ç›®å½•å­˜åœ¨ï¼ˆå…œåº•å¤„ç†ï¼‰
        new_path_dir = os.path.dirname(new_path)
        if not os.path.exists(new_path_dir):
            os.makedirs(new_path_dir)
        work_part.SaveAs(new_path)
        return True
    except Exception as e:
        print(f"ä¿å­˜éƒ¨ä»¶å¤±è´¥: {str(e)}")
        return False

# ==================================================================================
# 2. åˆ€è½¨ç”Ÿæˆå™¨ç±»
# ==================================================================================
class ToolpathGeneratorMacro:
    """åˆ€è½¨ç”Ÿæˆå™¨ - ä¸ºCAMæ“ä½œç”Ÿæˆåˆ€è½¨"""
    def __init__(self, session, work_part):
        self.session = session
        self.work_part = work_part
        self.success_count = 0
        self.failed_count = 0
        self.test_results = []

    def print_log(self, message, level="INFO"):
        """æ‰“å°å¸¦æ—¶é—´æˆ³çš„æ—¥å¿—ï¼ˆæ§åˆ¶å°è¾“å‡ºï¼Œæ›¿ä»£ListingWindowï¼‰"""
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        level_symbols = {
            "INFO": "â„¹ï¸", "SUCCESS": "âœ…", "ERROR": "âŒ",
            "WARN": "âš ï¸", "DEBUG": "ğŸ”", "START": "ğŸš€", "END": "ğŸ"
        }
        symbol = level_symbols.get(level, "â€¢")
        log_msg = f"[{timestamp}] {symbol} {message}"
        print(log_msg, flush=True)

    def print_separator(self, char="=", length=60):
        sep = char * length
        print(sep, flush=True)

    def print_header(self, title):
        self.print_separator()
        self.print_log(f"  {title}", "START")
        self.print_separator()

    def switch_to_manufacturing(self):
        try:
            self.session.ApplicationSwitchImmediate("UG_APP_MANUFACTURING")
            self.print_log("å·²åˆ‡æ¢åˆ°åŠ å·¥ç¯å¢ƒ", "SUCCESS")
        except Exception as e:
            self.print_log(f"åˆ‡æ¢åŠ å·¥ç¯å¢ƒè­¦å‘Š: {e}", "WARN")

    def get_all_operations(self):
        operations = []
        try:
            if self.work_part.CAMSetup is None:
                self.print_log("å½“å‰éƒ¨ä»¶æ— CAMåŠ å·¥ç¯å¢ƒ", "WARN")
                return operations
            for operation in self.work_part.CAMSetup.CAMOperationCollection:
                operations.append(operation)
            self.print_log(f"æ‰¾åˆ° {len(operations)} ä¸ªCAMæ“ä½œ", "INFO")
        except Exception as e:
            self.print_log(f"è·å–æ“ä½œåˆ—è¡¨å¤±è´¥: {e}", "ERROR")
        return operations

    def generate_toolpath(self, operation):
        op_name = operation.Name
        try:
            # è®¾ç½®Undoæ ‡è®°ï¼Œè™½ç„¶åé¢ä¸ä¸€å®šä¼šæ’¤é”€ï¼Œä½†ç”±äºAPIè¦æ±‚
            mark_id = self.session.SetUndoMark(
                NXOpen.Session.MarkVisibility.Visible,
                f"Generate Tool Path - {op_name}"
            )
            objects = [NXOpen.CAM.CAMObject.Null] * 1
            objects[0] = operation
            self.work_part.CAMSetup.GenerateToolPath(objects)
            self.print_log(f"åˆ€è½¨ç”Ÿæˆå®Œæˆ: {op_name}", "SUCCESS")
            self.success_count += 1
            self.test_results.append({
                "name": op_name, "status": "Success", "message": "åˆ€è½¨ç”ŸæˆæˆåŠŸ"
            })
            return True
        except Exception as e:
            err_msg = f"åˆ€è½¨ç”Ÿæˆå¤±è´¥ - {op_name}: {e}"
            self.print_log(err_msg, "ERROR")
            self.failed_count += 1
            self.test_results.append({
                "name": op_name, "status": "Failed", "error": str(e)
            })
            return False

    def generate_all_toolpaths(self):
        self.print_header("NX CAM åˆ€è½¨ç”Ÿæˆæµç¨‹")
        self.switch_to_manufacturing()
        operations = self.get_all_operations()

        if not operations:
            self.print_log("æ²¡æœ‰æ‰¾åˆ°ä»»ä½•CAMæ“ä½œ", "WARN")
            return

        # åˆ—å‡ºæ‰€æœ‰æ“ä½œ
        for i, op in enumerate(operations, 1):
            self.print_log(f"  {i}. {op.Name}")

        self.print_separator("-")
        self.print_log("å¼€å§‹ç”Ÿæˆåˆ€è½¨...", "START")
        for operation in operations:
            self.generate_toolpath(operation)

        self.print_summary()

    def print_summary(self):
        total = self.success_count + self.failed_count
        success_rate = (self.success_count / total * 100) if total > 0 else 0
        self.print_separator("=")
        summary = f"""
  åˆ€è½¨ç”Ÿæˆæ±‡æ€»
  ----------------------------------------
  æ€»æ“ä½œæ•°:   {total}
  æˆåŠŸ:       {self.success_count} âœ…
  å¤±è´¥:       {self.failed_count} âŒ
  æˆåŠŸç‡:     {success_rate:.1f}%
        """.strip()
        print(summary, flush=True)
        self.print_separator("=")

# ==================================================================================
# 3. åŠ ç‰¹æ—çº¿ç¨‹ç±» (æ¨¡æ‹ŸæŒ‰é”®å¤„ç†å¼¹çª—)
# ==================================================================================
class EnterSpammer(threading.Thread):
    def __init__(self, key_interval=0.3):
        super().__init__()
        self._stop_event = threading.Event()
        self.daemon = True
        self.key_interval = key_interval

    def run(self):
        user32 = ctypes.windll.user32
        VK_RETURN = 0x0D
        counter = 0
        while not self._stop_event.is_set():
            try:
                user32.keybd_event(VK_RETURN, 0, 0, 0)
                user32.keybd_event(VK_RETURN, 0, 2, 0)
                counter += 1
                time.sleep(self.key_interval)
            except Exception as e:
                # æ•è·æŒ‰é”®å¼‚å¸¸ï¼Œé¿å…çº¿ç¨‹å´©æºƒ
                time.sleep(self.key_interval)
                continue

    def stop(self):
        self._stop_event.set()

# ==================================================================================
# 4. è¿‡åˆ‡æ£€æŸ¥æ ¸å¿ƒæµç¨‹ï¼ˆä¿ç•™æ–‡ä»¶è¾“å‡ºï¼Œç§»é™¤çª—å£äº¤äº’ï¼‰
# ==================================================================================
def process_gauge_check(config):
    """æ‰§è¡Œè¿‡åˆ‡æ£€æŸ¥æµç¨‹"""
    session = NXOpen.Session.GetSession()
    workPart = session.Parts.Work
    if workPart is None:
        print("è¿‡åˆ‡æ£€æŸ¥å¤±è´¥ï¼šå½“å‰æ— å·¥ä½œéƒ¨ä»¶")
        return False

    # åˆ‡æ¢åŠ å·¥æ¨¡å—
    try:
        session.ApplicationSwitchImmediate("UG_APP_MANUFACTURING")
        uf = NXOpen.UF.UFSession.GetUFSession()
        uf.Cam.InitSession()
    except Exception as e:
        print(f"åˆ‡æ¢åŠ å·¥æ¨¡å—å¤±è´¥ï¼Œä½†ç»§ç»­æ‰§è¡Œè¿‡åˆ‡æ£€æŸ¥: {e}")

    # åˆå§‹åŒ–è¾“å‡ºçª—å£ï¼ˆä»…ç”¨äºå†™å…¥TXTæ–‡ä»¶ï¼Œæ— çª—å£äº¤äº’ï¼‰
    list_window = session.ListingWindow
    try:
        # ç¡®ä¿TXTæ–‡ä»¶ç›®å½•å­˜åœ¨ï¼ˆå…œåº•å¤„ç†ï¼‰
        txt_dir = os.path.dirname(config['OUTPUT_PATH'])
        if not os.path.exists(txt_dir):
            os.makedirs(txt_dir)
        list_window.SelectDevice(NXOpen.ListingWindow.DeviceType.File, config['OUTPUT_PATH'])
        list_window.Open()
    except Exception as e:
        print(f"åˆå§‹åŒ–è¾“å‡ºæ–‡ä»¶å¤±è´¥: {e}")
        return False

    infoTool = session.Information
    camSetup = workPart.CAMSetup
    if camSetup is None:
        print("è¿‡åˆ‡æ£€æŸ¥å¤±è´¥ï¼šå½“å‰éƒ¨ä»¶æ— CAMSetup")
        list_window.Close()
        return False

    # è·å–å·¥åºåˆ—è¡¨
    try:
        operations = [op for op in camSetup.CAMOperationCollection]
    except Exception as e:
        print(f"è·å–å·¥åºåˆ—è¡¨å¤±è´¥: {e}")
        list_window.Close()
        return False

    if not operations:
        print("è¿‡åˆ‡æ£€æŸ¥ï¼šæ— å·¥åºå¯å¤„ç†")
        list_window.Close()
        return True

    # å¯åŠ¨åŠ ç‰¹æ—çº¿ç¨‹
    spammer = EnterSpammer(config['KEY_INTERVAL'])
    spammer.start()

    count_success = 0
    try:
        # éå†å·¥åºæ‰§è¡Œè¿‡åˆ‡æ£€æŸ¥
        for i, op in enumerate(operations):
            op_name = op.Name
            op_type = str(type(op))

            # è¿‡æ»¤é’»å­”ç±»å·¥åº
            if any(x in op_type for x in ("Drill", "Hole", "Point")):
                continue

            msg = f"å‡†å¤‡å¤„ç†ç¬¬ {i+1} ä¸ªå·¥åº: [{op_name}]"
            list_window.WriteLine(msg)  # å†™å…¥æ–‡ä»¶ï¼Œä¿ç•™
            print(msg)  # æ§åˆ¶å°åŒæ­¥è¾“å‡º

            try:
                camSetup.GougeCheck([op], False)
                infoTool.DisplayCamObjectsDetails([op])

                success_msg = f"âœ… æˆåŠŸ: {op_name}"
                list_window.WriteLine(success_msg)
                print(success_msg)
                count_success += 1
            except Exception as e:
                err_msg = f"âŒ å¤±è´¥ {op_name}: {str(e)}"
                list_window.WriteLine(err_msg)
                print(err_msg)

            time.sleep(config['BETWEEN_OPS'])

        # ä¿å­˜è¿‡åˆ‡æ£€æŸ¥åçš„éƒ¨ä»¶ï¼ˆä»…ä¿ç•™æ­¤æ–‡ä»¶ï¼‰
        save_part(config['GAUGE_CHECK_SAVE_PATH'])

        # å†™å…¥å®Œæˆä¿¡æ¯
        finish_msg = f"è¿‡åˆ‡æ£€æŸ¥å®Œæˆã€‚æˆåŠŸ: {count_success} ä¸ª"
        list_window.WriteLine(finish_msg)
        print(finish_msg)

        return True
    finally:
        # åœæ­¢åŠ ç‰¹æ—çº¿ç¨‹
        spammer.stop()
        spammer.join()
        list_window.Close()  # å…³é—­æ–‡ä»¶å†™å…¥

# ==================================================================================
# 5. NXå·¥åºå‚æ•°å¯¼å‡ºå™¨ï¼ˆå®Œå…¨ç§»é™¤ListingWindowï¼‰
# ==================================================================================
class NXOperationParamExporter:
    """NXå·¥åºå‚æ•°å¯¼å‡ºå™¨ï¼Œç”¨äºæ‰¹é‡æ‰«æå¹¶æ±‡æ€»æ‰€æœ‰å·¥åºå‚æ•°åˆ°å•ä¸ªJSONæ–‡ä»¶"""

    def __init__(self, session=None, work_part=None, config=None):
        self.theSession = session or NXOpen.Session.GetSession()
        self.theUFSession = NXOpen.UF.UFSession.GetUFSession()
        self.workPart = work_part or self.theSession.Parts.Work
        self.config = config or {}  # ç§»é™¤lwç›¸å…³åˆå§‹åŒ–
        
        self.param_dictionary = self._get_param_dictionary()
        self.summary_data = {}
        self.batch_timestamp = time.strftime("%Y%m%d_%H%M%S", time.localtime())
        self.success_count = 0
        self.fail_count = 0

    def _get_param_dictionary(self):
        return {
            6:  "Display Tool Options",
            11: "Feed Engage",
            12: "Feed Retract",
            17: "Feed Cut",
            21: "Region Connection",
            24: "Corner Control Method",
            30: "Return Method",
            37: "Boundary Tolerances",
            42: "Min Clearance",
            45: "Start Method",
            49: "Gohome Method",
            55: "Motion Output Type",
            58: "NURBS Angular Tolerance",
            105: "Template Type",
            106: "Template Subtype",
            107: "Post Commands",
            116: "Suppressed",
            124: "Toolpath Time",
            125: "Toolpath Length",
            141: "Split Toolpath by Holder",
            142: "Toolpath Cutting Time",
            143: "Toolpath Cutting Length",
            148: "Postprocessor Cutting Time",
            153: "Template Class",
            154: "Template Subclass",
            221: "Clearance Plane Usage",
            1107: "Tool Adjust Reg Toggle",
            3010: "Hole Geometry",
            4005: "Spindle RPM",
            4013: "Spindle RPM Toggle",
            7210: "Last Tool Diameter",
            8212: "Tool Axis Tilt Data"
        }

    def init_environment(self):
        """åˆå§‹åŒ–ç¯å¢ƒï¼ˆæ§åˆ¶å°è¾“å‡ºï¼Œæ›¿ä»£çª—å£ï¼‰"""
        print("\næ­£åœ¨å¯åŠ¨å·¥åºå‚æ•°å¯¼å‡ºç¨‹åº...")

    def get_all_operations(self):
        """è·å–æ‰€æœ‰CAMå·¥åºï¼ˆç§»é™¤çª—å£è¾“å‡ºï¼Œæ”¹ä¸ºæ§åˆ¶å°ï¼‰"""
        camSetup = self.workPart.CAMSetup
        if camSetup is None:
            print("\nã€é”™è¯¯ã€‘å½“å‰éƒ¨ä»¶ä¸­æœªæ£€æµ‹åˆ°CAMåŠ å·¥ç¯å¢ƒï¼")
            return []

        opCollection = camSetup.CAMOperationCollection
        operations = [op for op in opCollection]
        if not operations:
            print("\nã€é”™è¯¯ã€‘å½“å‰éƒ¨ä»¶ä¸­æœªæ£€æµ‹åˆ°ä»»ä½•å·¥åºï¼")
        else:
            print(f"\næ£€æµ‹åˆ° {len(operations)} ä¸ªå·¥åºï¼Œå¼€å§‹æ‰¹é‡æ‰«æå‚æ•°...")
        return operations

    def scan_operation_params(self, op):
        """æ‰«æå•ä¸ªå·¥åºçš„å‚æ•°"""
        obj_tag = op.Tag
        collected_params = []
        # æ‰«æå‚æ•°èŒƒå›´ï¼š1-10000
        for index in range(1, 10000):
            val = None
            val_type = "Unknown"
            # Try Double
            if val is None:
                try:
                    val = self.theUFSession.Param.AskDoubleValue(obj_tag, index)
                    val_type = "Double"
                    if val == 0.0 and index not in self.param_dictionary:
                        val = None
                except:
                    pass
            # Try Int
            if val is None:
                try:
                    val = self.theUFSession.Param.AskIntValue(obj_tag, index)
                    val_type = "Int"
                    if val == 0 and index not in self.param_dictionary:
                        val = None
                except:
                    pass
            # Try String
            if val is None:
                try:
                    val = self.theUFSession.Param.AskStringValue(obj_tag, index)
                    val_type = "String"
                    if val == "":
                        val = None
                except:
                    pass
            # Try Tag
            if val is None:
                try:
                    val = self.theUFSession.Param.AskTagValue(obj_tag, index)
                    val_type = "Tag"
                    if val == NXOpen.Tag.Null:
                        val = None
                except:
                    pass

            if val is not None:
                display_name = self.param_dictionary.get(index, f"UNKNOWN_ID_{index}")
                collected_params.append({
                    "id": index,
                    "display_name": display_name,
                    "type": val_type,
                    "value": val
                })
        return collected_params

    def build_summary_data(self):
        """æ„å»ºæ±‡æ€»æ•°æ®"""
        nx_version = "NX"
        try:
            nx_version = self.theSession.EnvironmentInformation.Version
        except:
            pass

        self.summary_data = {
            "batch_meta": {
                "export_time": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
                "nx_version": nx_version,
                "part_name": self.workPart.Name if self.workPart else "æœªçŸ¥éƒ¨ä»¶",
                "total_operations": 0,
                "batch_timestamp": self.batch_timestamp,
                "success_operations": 0,
                "fail_operations": 0
            },
            "operations": []
        }

    def process_operations(self, operations):
        """å¤„ç†æ‰€æœ‰å·¥åºçš„å‚æ•°"""
        self.summary_data["batch_meta"]["total_operations"] = len(operations)
        for idx, op in enumerate(operations, 1):
            op_name = op.Name
            print(f"\n[{idx}/{len(operations)}] æ­£åœ¨å¤„ç†å·¥åºï¼š{op_name}")
            try:
                collected_params = self.scan_operation_params(op)
                self.summary_data["operations"].append({
                    "operation_name": op_name,
                    "operation_type": type(op).__name__,
                    "total_params": len(collected_params),
                    "parameters": collected_params,
                    "status": "success"
                })
                self.success_count += 1
                self.summary_data["batch_meta"]["success_operations"] = self.success_count
            except Exception as e:
                self.summary_data["operations"].append({
                    "operation_name": op_name,
                    "status": "failed",
                    "error_message": str(e)
                })
                self.fail_count += 1
                self.summary_data["batch_meta"]["fail_operations"] = self.fail_count
                print(f"    ã€å¤„ç†å¤±è´¥ã€‘: {str(e)}")

    def save_summary_file(self):
        """ä¿å­˜JSONæ–‡ä»¶"""
        base_dir = self.config.get('JSON_EXPORT_BASE_DIR', os.getcwd())
        if not os.path.exists(base_dir):
            try:
                os.makedirs(base_dir)
            except OSError as e:
                print(f"\nã€é”™è¯¯ã€‘æ— æ³•åˆ›å»ºç›®å½• {base_dir}ï¼š{str(e)}")
                return False

        full_path = self.config.get('JSON_FULL_PATH', os.path.join(base_dir, "data.json"))
        try:
            with open(full_path, "w", encoding='utf-8') as f:
                json.dump(self.summary_data, f, indent=4, ensure_ascii=False)
            print(f"\næˆåŠŸä¿å­˜JSONæ–‡ä»¶: {full_path}")
            return True
        except Exception as e:
            print(f"\nã€ä¿å­˜å¤±è´¥ã€‘: {str(e)}")
            return False

    def run(self):
        """æ‰§è¡Œå‚æ•°å¯¼å‡ºæµç¨‹"""
        self.init_environment()
        operations = self.get_all_operations()
        if not operations:
            return
        self.build_summary_data()
        self.process_operations(operations)
        self.save_summary_file()

# ==================================================================================
# 6. Excel æŠ¥å‘Šç”Ÿæˆå™¨ (æ•°æ®å¤„ç†é€»è¾‘)
# ==================================================================================
class ExcelReportGenerator:
    """å¤„ç†æ–‡æœ¬å’ŒJSONæ•°æ®ï¼Œç”ŸæˆExcelæŠ¥å‘Š"""
    
    @staticmethod
    def split_txt_by_generated_on(file_path):
        blocks = []
        current_block = []
        pattern = re.compile(r'\s*GENERATED\s+ON\s*', re.IGNORECASE)

        if not os.path.exists(file_path):
            print(f"ExcelGenerator: TXTæ–‡ä»¶ä¸å­˜åœ¨ {file_path}")
            return []

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line_stripped = line.rstrip('\n')
                    if pattern.search(line_stripped):
                        if current_block:
                            blocks.append(current_block)
                        current_block = [line_stripped]
                    else:
                        current_block.append(line_stripped)
                if current_block:
                    blocks.append(current_block)
            return blocks
        except Exception as e:
            print(f"ExcelGenerator: è¯»å–æ–‡ä»¶é”™è¯¯ {e}")
            return []

    @staticmethod
    def extract_process_info(block):
        # 1. æå–å·¥åºå
        process_pattern = re.compile(r'(Operation Name)\s*[:ï¼š\s]\s*(.+)', re.IGNORECASE)
        # 2. æå–è¿‡åˆ‡çŠ¶æ€
        gouge_pattern = re.compile(r'(Gouge Check Status)\s*[:ï¼š\s]\s*([^ï¼›ï¼Œã€‚\n]+)', re.IGNORECASE)

        process_name = "æœªçŸ¥å·¥åº"
        gouge_status = "None"

        for line in block:
            if process_name == "æœªçŸ¥å·¥åº":
                process_match = process_pattern.search(line)
                if process_match:
                    process_name = process_match.group(2).strip()
            if gouge_status == "None":
                gouge_match = gouge_pattern.search(line)
                if gouge_match:
                    gouge_status = gouge_match.group(2).strip()
            if process_name != "æœªçŸ¥å·¥åº" and gouge_status != "None":
                break
        return process_name, gouge_status

    @staticmethod
    def extract_toolpath_time_from_json(json_path):
        toolpath_times = []
        if not os.path.exists(json_path):
            print(f"ExcelGenerator: JSONæ–‡ä»¶ä¸å­˜åœ¨ {json_path}")
            return toolpath_times
        
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            operations = data.get('operations', [])
            for op in operations:
                toolpath_time = 0.0
                parameters = op.get('parameters', [])
                for param in parameters:
                    display_name = param.get('display_name', '').strip().lower()
                    if 'toolpath time' in display_name or 'åŠ å·¥æ—¶é—´' in display_name:
                        try:
                            toolpath_time = float(param.get('value', 0.0))
                        except:
                            toolpath_time = 0.0
                        break
                toolpath_times.append(toolpath_time)
            print(f"ä»JSONä¸­æå–åˆ° {len(toolpath_times)} ä¸ªToolpath Timeå€¼")
            return toolpath_times
        except Exception as e:
            print(f"ExcelGenerator: JSONè§£æé”™è¯¯ {e}")
            return toolpath_times

    @staticmethod
    def write_to_excel(excel_path, process_names, gouge_statuses, toolpath_statuses, part_file_path):
        from openpyxl.utils import get_column_letter  # å±€éƒ¨å¯¼å…¥é¿å…å½±å“é¡¶éƒ¨
        try:
            # ç¡®ä¿Excelç›®å½•å­˜åœ¨ï¼ˆå…œåº•å¤„ç†ï¼‰
            excel_dir = os.path.dirname(excel_path)
            if not os.path.exists(excel_dir):
                os.makedirs(excel_dir)

            if os.path.exists(excel_path):
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "CAMè¿‡åˆ‡æ£€æŸ¥ç»“æœ"
                ws.cell(row=1, column=1, value="æ–‡ä»¶è·¯å¾„")    # A1: æ–°å¢
                ws.cell(row=1, column=2, value="å·¥åºå")
                ws.cell(row=1, column=3, value="æ˜¯å¦è¿‡åˆ‡")
                ws.cell(row=1, column=4, value="æ˜¯å¦ç”Ÿæˆåˆ€è·¯")
                print(f"å·²åˆ›å»ºæ–°Excelæ–‡ä»¶ï¼š{excel_path}")

            # ç¡®ä¿è¡¨å¤´å­˜åœ¨
            if ws.cell(row=1, column=1).value != "æ–‡ä»¶è·¯å¾„":
                ws.cell(row=1, column=1, value="æ–‡ä»¶è·¯å¾„")

            process_col = 2
            gouge_col = 3
            toolpath_col = 4
            start_row = 2

            # å†™å…¥æ•°æ®
            max_len = max(len(process_names), len(gouge_statuses), len(toolpath_statuses))
            for idx in range(max_len):
                # å†™å…¥æ–‡ä»¶è·¯å¾„
                ws.cell(row=start_row + idx, column=1, value=part_file_path)
                
                if idx < len(process_names):
                    ws.cell(row=start_row + idx, column=process_col, value=process_names[idx])
                if idx < len(gouge_statuses):
                    ws.cell(row=start_row + idx, column=gouge_col, value=gouge_statuses[idx])
                if idx < len(toolpath_statuses):
                    ws.cell(row=start_row + idx, column=toolpath_col, value=toolpath_statuses[idx])
                else:
                    ws.cell(row=start_row + idx, column=toolpath_col, value="æœªçŸ¥")

            # è‡ªåŠ¨è°ƒæ•´åˆ—å®½
            for column_cells in ws.columns:
                length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
                # é€‚å½“å¢åŠ ä¸€ç‚¹å®½åº¦ä½œä¸ºç¼“å†²ï¼Œå¹¶è®¾ç½®ä¸Šé™é˜²æ­¢è¿‡å®½
                adjusted_width = min((length + 2) * 1.2, 50) 
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

            wb.save(excel_path)
            print(f"ExcelæŠ¥å‘Šç”ŸæˆæˆåŠŸï¼å…±å†™å…¥ {max_len} æ¡æ•°æ®ã€‚")
            print(f"è·¯å¾„: {excel_path}")
        except PermissionError:
            print("é”™è¯¯ï¼šExcelæ–‡ä»¶å·²è¢«å ç”¨ï¼Œè¯·å…³é—­åé‡è¯•ã€‚")
        except Exception as e:
            print(f"Excelå†™å…¥å¤±è´¥ï¼š{e}")
            traceback.print_exc()

    @classmethod
    def run_report_generation(cls, config):
        print("\n===== å¼€å§‹ç”ŸæˆExcelæŠ¥å‘Š =====")
        
        # 1. å¤„ç†TXT
        txt_blocks = cls.split_txt_by_generated_on(config['OUTPUT_PATH'])
        if not txt_blocks:
            print("æç¤º: æœªä»TXTä¸­æå–åˆ°æœ‰æ•ˆå—ï¼Œè·³è¿‡")
            return

        process_info = [cls.extract_process_info(block) for block in txt_blocks]
        process_names = [info[0] for info in process_info]
        gouge_statuses = [info[1] for info in process_info]

        # 2. å¤„ç†JSON
        toolpath_times = cls.extract_toolpath_time_from_json(config['JSON_FULL_PATH'])

        # 3. å¯¹é½æ•°æ® (ä»¥JSONä¸ºå‡†)
        min_len = len(toolpath_times)
        if len(process_names) > min_len:
            print(f"ExcelGenerator: TXTæ•°æ®({len(process_names)}) å¤šäº JSONæ•°æ®({min_len})ï¼Œæ­£åœ¨æˆªæ–­...")
            process_names = process_names[:min_len]
            gouge_statuses = gouge_statuses[:min_len]
        elif len(toolpath_times) > len(process_names):
            print(f"ExcelGenerator: JSONæ•°æ®({len(toolpath_times)}) å¤šäº TXTæ•°æ®({len(process_names)})ï¼Œéƒ¨åˆ†è¡Œå°†ä¸ºç©º")

        # 4. ç”Ÿæˆåˆ€è·¯ç”ŸæˆçŠ¶æ€
        toolpath_statuses = ["æ˜¯" if t > 0 else "å¦" for t in toolpath_times]

        # 5. å†™å…¥Excel
        cls.write_to_excel(config['EXCEL_REPORT_PATH'], process_names, gouge_statuses, toolpath_statuses, config['PART_PATH'])

# ==================================================================================
# 7. ä¸»å·¥ä½œæµï¼ˆæ ¸å¿ƒï¼šä»…æ¥æ”¶prtè·¯å¾„å’Œæ ¹æ–‡ä»¶å¤¹è·¯å¾„ï¼Œæ–‡ä»¶åç§°æ‹¼æ¥PRTåç§°ï¼‰
# ==================================================================================
def main(
    part_path: str,
    root_dir: str
):
    """æ•´åˆä¸»æµç¨‹ï¼šåˆ€è½¨ -> è¿‡åˆ‡ -> JSON -> Excel
    Args:
        part_path: PRTéƒ¨ä»¶æ–‡ä»¶çš„å®Œæ•´è·¯å¾„
        root_dir: æ ¹æ–‡ä»¶å¤¹è·¯å¾„ï¼Œè„šæœ¬ä¼šåœ¨è¯¥è·¯å¾„ä¸‹è‡ªåŠ¨åˆ›å»ºexcel/txt/json/prtå­æ–‡ä»¶å¤¹
    """
    # 1. å›ºå®šé…ç½®ï¼ˆä¿ç•™ä¸å˜çš„å‚æ•°ï¼Œæ— éœ€å¤–éƒ¨ä¼ å…¥ï¼‰
    fixed_config = {
        "AUTO_SAVE_TOOLPATH": False,
        "HEART_BEAT_EVERY": 10,
        "KEY_INTERVAL": 0.3,
        "BETWEEN_OPS": 0.5
    }

    # 2. è‡ªåŠ¨åˆ›å»ºæ ¹ç›®å½•ï¼ˆå¦‚æœä¸å­˜åœ¨ï¼‰
    if not os.path.exists(root_dir):
        os.makedirs(root_dir)
        print(f"å·²åˆ›å»ºæ ¹æ–‡ä»¶å¤¹: {root_dir}")

    # 3. è‡ªåŠ¨åˆ›å»ºå­æ–‡ä»¶å¤¹
    sub_dirs = ["excel", "txt", "json", "prt"]
    sub_dir_paths = {}
    for sub_dir in sub_dirs:
        sub_dir_path = os.path.join(root_dir, sub_dir)
        if not os.path.exists(sub_dir_path):
            os.makedirs(sub_dir_path)
            print(f"å·²åˆ›å»ºå­æ–‡ä»¶å¤¹: {sub_dir_path}")
        sub_dir_paths[sub_dir] = sub_dir_path

    # 4. æå–PRTæ–‡ä»¶åï¼ˆæ ¸å¿ƒï¼šåªæå–ä¸€æ¬¡ï¼Œå…¨ç¨‹å¤ç”¨ï¼‰
    part_name = os.path.splitext(os.path.basename(part_path))[0]

    # 5. è‡ªåŠ¨ç”Ÿæˆå„è¾“å‡ºæ–‡ä»¶çš„å®Œæ•´è·¯å¾„ï¼ˆæ‹¼æ¥PRTåç§°ï¼Œä¿è¯ä¸€è‡´æ€§ï¼‰
    dynamic_path_config = {
        "PART_PATH": part_path,
        # PRTå­æ–‡ä»¶å¤¹ï¼šPRTå + _guoqie.prt
        "GAUGE_CHECK_SAVE_PATH": os.path.join(sub_dir_paths["prt"], f"{part_name}_guoqie.prt"),
        # TXTå­æ–‡ä»¶å¤¹ï¼šPRTå + _Feature_Info.txt
        "OUTPUT_PATH": os.path.join(sub_dir_paths["txt"], f"{part_name}_Feature_Info.txt"),
        # JSONå­æ–‡ä»¶å¤¹ï¼šPRTå + _data.json
        "JSON_EXPORT_BASE_DIR": sub_dir_paths["json"],
        "JSON_FULL_PATH": os.path.join(sub_dir_paths["json"], f"{part_name}_data.json"),
        # Excelå­æ–‡ä»¶å¤¹ï¼šPRTå + _CAMè¿‡åˆ‡æ£€æŸ¥.xlsx
        "EXCEL_REPORT_PATH": os.path.join(sub_dir_paths["excel"], f"{part_name}_CAMè¿‡åˆ‡æ£€æŸ¥.xlsx")
    }

    # 6. åˆå¹¶é…ç½®ï¼ˆåŠ¨æ€è·¯å¾„ + å›ºå®šå‚æ•°ï¼‰
    config = {**dynamic_path_config, **fixed_config}

    # 7. å¿…ä¼ å‚æ•°æ ¡éªŒ
    if not os.path.exists(part_path):
        raise ValueError(f"é”™è¯¯ï¼šPRTæ–‡ä»¶ä¸å­˜åœ¨ -> {part_path}")

    # åç»­æ ¸å¿ƒä¸šåŠ¡é€»è¾‘
    session = NXOpen.Session.GetSession()

    # æ‰“å¼€éƒ¨ä»¶
    print(f"\næ­£åœ¨æ‰“å¼€éƒ¨ä»¶: {config['PART_PATH']}")
    try:
        base_part, load_status = session.Parts.OpenBaseDisplay(config['PART_PATH'])
    except Exception as e:
        print(f"æ‰“å¼€éƒ¨ä»¶å¤±è´¥: {e}")
        return False

    if not check_part_load_status(load_status):
        print("âŒ æ‰“å¼€éƒ¨ä»¶å¤±è´¥")
        return False
    work_part = session.Parts.Work

    # ç”Ÿæˆåˆ€è½¨
    generator = ToolpathGeneratorMacro(session, work_part)
    generator.generate_all_toolpaths()
    generator.print_log("åˆ€è½¨ç”Ÿæˆæµç¨‹ç»“æŸ", "END")

    # æ‰§è¡Œè¿‡åˆ‡æ£€æŸ¥ï¼ˆç”ŸæˆTXTï¼‰
    print("\n===== å¼€å§‹æ‰§è¡Œè¿‡åˆ‡æ£€æŸ¥ =====")
    process_gauge_check(config)

    # å¯¼å‡ºå·¥åºå‚æ•°ä¸ºJSON
    print("\n===== å¼€å§‹å¯¼å‡ºå·¥åºå‚æ•°ä¸ºJSON =====")
    exporter = NXOperationParamExporter(session, work_part, config)
    exporter.run()

    # ç”ŸæˆExcelæŠ¥å‘Š
    ExcelReportGenerator.run_report_generation(config)

    print("\n===== æ‰€æœ‰æµç¨‹æ‰§è¡Œå®Œæˆ =====")
    print(f"è¾“å‡ºæ–‡ä»¶æ ¹ç›®å½•: {root_dir}")
    return True

# ==================================================================================
# ä¸»ç¨‹åºå…¥å£ï¼ˆæµ‹è¯•ç”¨ï¼‰
# ==================================================================================
if __name__ == "__main__":
    # æœ¬åœ°æµ‹è¯•æ—¶æ‰‹åŠ¨ä¼ å…¥ä¸¤ä¸ªå‚æ•°ï¼šPRTè·¯å¾„ + æ ¹æ–‡ä»¶å¤¹è·¯å¾„
    try:
        part_path = r"C:\z_èµ„æ–™\é’‰é’‰ç¾¤\P4\DIE-05_dwg.prt"
        root_dir = r"C:\Users\Admin\Desktop\NX_CAM_Output"

        main(part_path, root_dir)
    except Exception as e:
        print(f"\nâŒ ç¨‹åºæ‰§è¡Œå¼‚å¸¸: {e}")
        traceback.print_exc()