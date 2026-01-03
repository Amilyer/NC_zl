import math
import os
import re
import traceback
from collections import defaultdict

import NXOpen
from openpyxl import Workbook
from openpyxl.utils import get_column_letter  # 用于将列索引转换为列字母（如12→L）


def select_red_straight_lines(workPart) -> bool:
    """
    选择当前工作部件中所有红色（颜色186）的实线直线，判定是否存在研磨符号（6短+1长红色单实线）
    核心：保留原长度获取逻辑（displayable_object.GetLength()），不修改筛选基础逻辑
    返回值：存在研磨符号返回True，不存在返回False
    """
    try:
        # 配置参数（可按需调整，不影响原长度获取逻辑）
        TARGET_COLOR = 186  # 红色颜色值
        SHORT_LINE_MIN_COUNT = 6  # 短实线最小数量
        LENGTH_RATIO = 3.0  # 长短线长度比例
        LEN_TOLERANCE = 0.1  # 同组线长度误差
        RATIO_TOLERANCE = 0.2  # 比例误差（3倍±20%）
        POINT_COINCIDE_TOLERANCE = 0.001  # 相邻验证阈值

        # --------------------------
        # 保留原函数核心逻辑：筛选红色实线直线（不修改任何代码）
        # --------------------------
        red_straight_lines = []
        # 新增：存储线的关键信息（供后续判定用，不影响原筛选逻辑）
        red_line_details = []

        # 遍历所有曲线
        for curve in workPart.Curves:
            # 检查是否为直线
            if isinstance(curve, NXOpen.Line):
                try:
                    # 获取直线的显示属性
                    displayable_object = curve
                    if hasattr(curve, 'GetDisplayableObject'):
                        displayable_object = curve.GetDisplayableObject()

                    # 获取颜色
                    color = displayable_object.Color

                    if color == TARGET_COLOR:
                        # 保留原长度获取逻辑：displayable_object.GetLength()
                        _len = displayable_object.GetLength()
                        red_straight_lines.append(curve)

                        # 新增：存储长度、端点信息
                        start_point = curve.StartPoint
                        end_point = curve.EndPoint
                        red_line_details.append({
                            "line_obj": curve,
                            "length": _len,  # 直接使用原逻辑获取的长度
                            "start": (start_point.X, start_point.Y, start_point.Z),
                            "end": (end_point.X, end_point.Y, end_point.Z)
                        })
                except Exception as e:
                    # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
                    exc_info = traceback.extract_tb(e.__traceback__)[-1]
                    err_msg = (f"[select_red_straight_lines-曲线处理] 异常：{str(e)} | "
                               f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
                    print(err_msg)
                    continue

        # 保留原函数的基础输出
        if not red_straight_lines:
            return False  # 无红色线，返回False

        # --------------------------
        # 新增：研磨符号判定逻辑
        # --------------------------
        # 1. 长度分组（按误差聚类）
        length_group = defaultdict(list)
        for line_info in red_line_details:
            target_len = line_info["length"]
            found = False
            # 查找相近长度的分组
            for exist_len in list(length_group.keys()):
                if abs(target_len - exist_len) <= LEN_TOLERANCE:
                    length_group[exist_len].append(line_info)
                    found = True
                    break
            if not found:
                length_group[target_len].append(line_info)

        # 2. 筛选短实线组（数量≥6条）
        short_groups = [
            (s_len, s_lines) for s_len, s_lines in length_group.items()
            if len(s_lines) >= SHORT_LINE_MIN_COUNT
        ]
        if not short_groups:
            return False

        # 3. 匹配3倍长度的长实线+相邻验证
        for short_len, short_lines in short_groups:
            # 计算长实线长度范围
            min_long_len = short_len * (LENGTH_RATIO - RATIO_TOLERANCE)
            max_long_len = short_len * (LENGTH_RATIO + RATIO_TOLERANCE)

            # 查找符合长度的长实线
            for long_len, long_lines in length_group.items():
                if min_long_len <= long_len <= max_long_len:
                    # 验证长实线与短实线是否相邻（端点重合）
                    short_points = [p for line in short_lines for p in [line["start"], line["end"]]]
                    for long_line in long_lines:
                        long_points = [long_line["start"], long_line["end"]]
                        for lp in long_points:
                            for sp in short_points:
                                if math.dist(lp, sp) <= POINT_COINCIDE_TOLERANCE:
                                    # 找到1个符合条件，立即返回True
                                    return True

        # 未找到任何符合条件的组合
        return False
    except Exception as e:
        # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
        exc_info = traceback.extract_tb(e.__traceback__)[-1]
        err_msg = (f"[select_red_straight_lines] 异常：{str(e)} | "
                   f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
        print(err_msg)
        return False


def contains_text(text: str, substr: str) -> bool:
    """判断文本是否包含指定子串（不区分大小写）"""
    try:
        if not text or not substr:
            return False
        return substr.upper() in text.upper()
    except Exception as e:
        # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
        exc_info = traceback.extract_tb(e.__traceback__)[-1]
        err_msg = (f"[contains_text] 异常：{str(e)} | "
                   f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
        print(err_msg)
        return False


def get_wire_cuttings(text_dict, WE):
    """根据加工对象获取线割工艺"""
    try:
        parts = {"上垫板": WE[1], "上夹板": WE[0], "止挡板": WE[2],
                 "脱料板": WE[0], "上脱板": WE[0], "下模板": WE[0],
                 "下脱板": WE[2], "抬料板": WE[2], "弹簧座": WE[1],
                 "定位块": WE[1]
                 }
        parts_1 = {
            f"{WE[0]}": ['入块', '刀口', '冲孔', '抽牙', '滑块', '连接块', '切冲冲头'],
            f"{WE[2]}": ['成型', '折弯', '推平', '翻边', '成型冲头'],
            f"{WE[1]}": ['导料块', '字印', '压毛边', '凸包']
        }
        for part in parts:
            if part in text_dict["加工对象"]:
                return parts[part]
        for part_1 in parts_1:
            for name in parts_1[part_1]:
                if name in text_dict["加工对象"]:
                    return part_1
        return None
    except Exception as e:
        # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
        exc_info = traceback.extract_tb(e.__traceback__)[-1]
        err_msg = (f"[get_wire_cuttings] 异常：{str(e)} | "
                   f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
        print(err_msg)
        return None


def get_sort_rule(_is):
    try:
        # 定义WE_TYPE到具体代号的映射：0=快走丝(W/C)，1=中走丝(W/Z)，其他=慢走丝(W/E)
        we_type_map = {0: "W/C", 1: "W/Z"}
        # 热处理
        if _is["is_PTR"]:
            # 常规热处理
            sort_list = ["S", "Z", "PTR", "M", "WE", "YM", "SS", "QC"]
            # 存在自找料且存在牙孔（存在放电加工工序）
            if _is["is_special"] and _is["is_edm"]:
                sort_list = ["S", "EDM", "PTR", "M", "WE", "YM", "SS", "QC"]

        else:
            # 非热处理
            # 常规非热处理
            sort_list = ["M", "S", "Z", "SS", "WE", "QC"]
            # 存在自找料
            if _is["is_special"]:
                # 硬料直接加工
                if _is["is_hard_material"]:
                    sort_list = ["WE", "SS", "EDM", "QC"]
                else:
                    sort_list = ["WE", "M", "S", "Z", "SS", "QC"]

        # 将sort_list中的WE替换为具体代号（根据WE_TYPE）
        if "WE" in sort_list:
            # 获取具体WE代号（默认慢走丝W/E）
            we_code = we_type_map.get(_is["WE_TYPE"], "W/E")
            # 替换所有WE为具体代号
            sort_list = [we_code if item == "WE" else item for item in sort_list]

        return sort_list

    except Exception as e:
        # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
        exc_info = traceback.extract_tb(e.__traceback__)[-1]
        err_msg = (f"[get_sort_rule] 异常：{str(e)} | "
                   f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
        print(err_msg)
        return []


def get_material(text):
    """提取材质（优先匹配最长材质名称）"""
    try:
        materials = ["45#", "A3", "CR12", "CR12MOV", "SKD11", "SKH-9", "DC53", "P20", "TOOLOX33", "TOOLOX44", "合金铜",
                     "T00L0X33", "T00L0X44"]
        _material = None
        length = 0
        for material in materials:
            # 统一转为小写（或大写）比较，不区分大小写
            if material.lower() in text.lower():
                now_len = len(material)
                # 最匹配材质（取最长匹配）
                if now_len > length:
                    _material = material
                    length = now_len
        if _material == "T00L0X33":
            _material = "TOOLOX33"
        if _material == "T00L0X44":
            _material = "TOOLOX44"
        return _material
    except Exception as e:
        # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
        exc_info = traceback.extract_tb(e.__traceback__)[-1]
        err_msg = (f"[get_material] 异常：{str(e)} | "
                   f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
        print(err_msg)
        return None


def get_we_craft(text, text_dict, WE):
    """获取线割具体工艺（慢丝/中丝/快丝）"""
    try:
        for we in WE:
            if we in text:
                # 加工说明中存在"慢走丝" 或者 "中走丝" 或者 "快走丝"
                return we
        # 根据零件名称判断
        _we = get_wire_cuttings(text_dict, WE)
        if _we:
            return _we
        else:
            return "快走丝"
    except Exception as e:
        # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
        exc_info = traceback.extract_tb(e.__traceback__)[-1]
        err_msg = (f"[get_we_craft] 异常：{str(e)} | "
                   f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
        print(err_msg)
        return "快走丝"


def get_craft(workPart, text_dict, num_text, judgement_M=False):
    """
    提取加工工艺和工艺属性
    Args:
        workPart: 当前工作部件
        text_dict： 加工说明字典
        num_text：纯数字文本-图纸尺寸
        judgement_M：是否需要通过零件名称判断是否需要研磨
    Returns:
        crafts: 工艺字典
        _is: 存在工艺字典
    """
    try:
        # 钻孔工艺关键词
        NC_keywords = ["钻", "攻", "铰", "穿线孔", "沉头", "合销", "螺丝", "弹簧孔", "镗", "<O>", "Φ", "精镗"]
        # 开粗&精铣工艺关键词
        S_keywords = ["铣", "削"]
        # 线割工艺关键词
        WE_keywords = ["割"]
        WE = ["慢走丝", "快走丝", "中走丝"]  # 线割具体工艺
        # 热处理工艺关键词
        PTR_keywords = ["HRC", "hrc", "真空", "热处理", "调质", "激光", "深冷"]
        EDM_keywords = ["螺纹", "螺丝", "止付螺丝", "起吊螺纹", "牙孔"]
        # C03、C04材质类型--硬料
        hard_materials = ["CR12", "CR12MOV", "SKD11", "SKH-51", "SKH-9", "DC53", "CR8", "7CrSiMnMov"]
        # 加工说明文本-str
        text = f"{text_dict}"
        # 涉及工艺列表
        crafts = {}
        _is = {"is_PTR": False, "is_edm": False, "is_special": False, "is_hard_material": False, "is_SS": False,
               "is_WE": False, "WE_TYPE": 0, "sided_num": "一面加工", "开粗与钻孔同面": False, "型面与钻孔同面": False}

        # 双面加工
        if ("正" in text and "背" in text) or "双面" in text:
            _is["sided_num"] = "两面加工"
        # 自找料
        if "自找料" in text:
            _is["is_special"] = True

        # 提取材质
        material = get_material(text_dict.get("尺寸", ""))
        # 判断是否为硬料
        if material:
            if contains_text(f"{hard_materials}", f"{material}"):
                _is["is_hard_material"] = True
        # 热处理工艺
        if any(contains_text(text, kw) for kw in PTR_keywords):
            _is["is_PTR"] = True
            crafts["热处理"] = "(0.0mm)"
        # 放电加工
        if any(contains_text(text, kw) for kw in EDM_keywords):
            _is["is_edm"] = True
            crafts["放电加工"] = "(0.0mm)"
        # 磨床工艺
        has_grinding = select_red_straight_lines(workPart)  # 是否存在研磨符号
        if has_grinding:
            try:
                max_value = max(float(num_str) for num_str in num_text)
                if max_value > 400:
                    crafts["大水磨"] = "(0.0mm)"
                else:
                    crafts["小磨床"] = "(0.0mm)"
            except Exception as e:
                # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
                exc_info = traceback.extract_tb(e.__traceback__)[-1]
                err_msg = (f"[get_craft-磨床类型判断] 异常：{str(e)} | "
                           f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
                print(err_msg)
                print("未提取到有效尺寸信息，无法判断大小磨床")
                crafts["小磨床"] = "(0.0mm)"  # 异常时默认磨床

            # 研磨工艺：
            if ("垫脚" in text_dict.get("加工对象", "") or "托板" in text_dict.get("加工对象", "")) and judgement_M:
                pass
            else:
                pattern = r'.*?拼.*?磨'
                if bool(re.search(pattern, text)):
                    crafts["大水磨"] = "(0.1mm)"

        """
        存在自找料且存在放电加工(存在热处理以及存在牙孔)--不再执行钻孔加工，而是执行放电加工
        """
        if not _is["is_special"]:
            # 钻孔工艺
            if any(contains_text(text, kw) for kw in NC_keywords):
                crafts["钻孔"] = '(0.0mm)'
        # 线割工艺
        if any(contains_text(text, kw) for kw in WE_keywords):
            # 加工说明中存在"割"
            we = get_we_craft(text, text_dict, WE)
            if we:
                _is["is_WE"] = True
                if we == "快走丝":
                    _is["WE_TYPE"] = 0
                    crafts["快走丝"] = "(0.0mm)"
                elif we == "中丝":
                    _is["WE_TYPE"] = 1
                    crafts["中走丝"] = "(0.0mm)"
                else:
                    _is["WE_TYPE"] = 2
                    crafts["慢走丝"] = "(0.0mm)"

        # NC开粗工艺$NC精铣工艺
        pattern = r'与.*?拼.*?铣'
        is_pattern = bool(re.search(pattern, text))
        is_ss = False
        is_s = False
        for key in text_dict:
            if key in ["加工对象", "重量", "尺寸"]:
                continue
            value = str(text_dict[key])
            if any(contains_text(value, kw) for kw in S_keywords):
                if is_s and is_ss:
                    break
                if "精" in value or "单+" in value:
                    is_ss = True
                    continue
                crafts["CNC开粗"] = '(侧面0.8mm 正面0.0mm)'
                is_s = True
        if is_ss:
            _is["is_SS"] = True
            if is_pattern:
                crafts["CNC精铣"] = '(0.05mm)'
            else:
                crafts["CNC精铣"] = '(0.0mm)'

        return crafts, _is
    except Exception as e:
        # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
        exc_info = traceback.extract_tb(e.__traceback__)[-1]
        err_msg = (f"[get_craft] 异常：{str(e)} | "
                   f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
        print(err_msg)
        return {}, {}


def get_num_notes(workPart):
    """
    获取图纸尺寸（纯数字注释）用于判断大小磨床
    Returns:
        纯数字注释列表
    """
    try:
        # 提取工作部件中的所有注释
        notes = list(workPart.Notes)
        # 纯数字校验正则：仅允许数字和1个小数点（避免多小数点无效格式）
        pure_num_pattern = re.compile(r'^[0-9]+(\.[0-9]+)?$')
        # 第一步：提取原始注释数据（仅保留纯数字文本）
        raw_data_list = []
        for idx, note in enumerate(notes, start=1):
            try:
                # 获取注释文本（NX注释文本为数组，拼接为字符串并去除空格）
                texts = note.GetText()
                text_str = " ".join(texts).strip().replace(' ', '')

                # 新增：纯数字校验，非纯数字直接跳过
                if pure_num_pattern.match(text_str):
                    raw_data_list.append(text_str)

            except Exception as e:
                # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
                exc_info = traceback.extract_tb(e.__traceback__)[-1]
                err_msg = (f"[get_num_notes-注释读取] 异常：{str(e)} | "
                           f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
                print(err_msg)
                continue

        return raw_data_list
    except Exception as e:
        # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
        exc_info = traceback.extract_tb(e.__traceback__)[-1]
        err_msg = (f"[get_num_notes] 异常：{str(e)} | "
                   f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
        print(err_msg)
        return []


def safe_origin(note):
    """尝试多种方式读取坐标"""
    try:
        if hasattr(note, "AnnotationOrigin"):
            o = note.AnnotationOrigin
            return (o.X, o.Y, o.Z)
        elif hasattr(note, "Origin"):
            o = note.Origin
            return (o.X, o.Y, o.Z)
        elif hasattr(note, "GetOrigin"):
            o = note.GetOrigin()
            return (o.X, o.Y, o.Z)
    except Exception as e:
        # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
        exc_info = traceback.extract_tb(e.__traceback__)[-1]
        err_msg = (f"[safe_origin] 异常：{str(e)} | "
                   f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
        print(err_msg)
        return (None, None, None)
    return (None, None, None)


def _format_text_data(format_data_list, entries):
    """格式化NX提取的原始注释文本"""
    # 保留"加工说明"条目原样
    if "加工说明" in entries["文本"]:
        format_data_list.append({"标题": entries})
        return format_data_list

    # 跳过纯字母/字母数字文本
    if entries["文本"].isalpha() or entries["文本"].isalnum():
        return format_data_list

    # 合并以逗号开头的文本
    if entries["文本"].startswith(','):
        if format_data_list:
            last_entry = format_data_list[-1]
            # 避免重复合并
            if entries["文本"][1:] in last_entry["文本"]:
                return format_data_list
            # 合并文本
            if last_entry["文本"].endswith(",") or last_entry["文本"].endswith("，"):
                last_entry["文本"] += entries["文本"][1:]
            else:
                last_entry["文本"] += entries["文本"]
            # 合并ID和坐标
            last_entry["ID"] += entries["ID"]
            last_entry["坐标"] += entries["坐标"]
            return format_data_list

    # 非合并场景，直接添加到列表
    format_data_list.append(entries)
    return format_data_list


def extract_and_process_notes(work_part):
    """从NX当前工作部件中提取所有注释，返回结构化注释字典"""

    # 提取原始注释数据
    raw_data_list = []
    note_list = list(work_part.Notes)

    # 排除非加工信息、去重、重排序
    note_str = []
    notes = []
    for note in note_list:
        text_str = " ".join(note.GetText()).strip()
        if text_str in note_str:  # 去重
            continue
        # 保留不满足删除条件的note及对应文本
        if not len(text_str) < 4:
            try:
                float(text_str)
                continue
            except:
                note_str.append(text_str)
                notes.append(note)
    if "加工说明" not in note_str[0]:
        notes = notes[::-1]
    for idx, note in enumerate(notes, start=1):
        try:
            # 获取注释文本
            texts = note.GetText()
            text_str = " ".join(texts).strip()
            is_size = re.search(r'(\d+(?:\.\d+)?)L\D*?(\d+(?:\.\d+)?)W\D*?(\d+(?:\.\d+)?)T', text_str)
            if "HRC" not in text_str and not is_size:
                text_str = text_str.replace(" ", "")
                # 匹配尺寸信息
            # 获取注释坐标
            x, y, z = safe_origin(note)
            # 获取注释ID
            try:
                journal_id = note.JournalIdentifier
            except:
                journal_id = "N/A"

            # 构建原始注释字典
            raw_entry = {
                "ID": [journal_id],
                "文本": text_str,
                "坐标": [f"({x:.3f}, {y:.3f}, {z:.3f})"] if x is not None else ["(None, None, None)"]
            }
            # 初步格式化（合并拆分注释）
            raw_data_list = _format_text_data(raw_data_list, raw_entry)

        except Exception as e:
            # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
            exc_info = traceback.extract_tb(e.__traceback__)[-1]
            err_msg = (f"[extract_and_process_notes-注释处理] 异常：{str(e)} | "
                       f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
            print(err_msg)

    # 结构化分类注释数据
    final_dict = {"其它未注解按3D涂色加工": []}

    size_pattern = r'(\d+(?:\.\d+)?)L\D*?(\d+(?:\.\d+)?)W\D*?(\d+(?:\.\d+)?)T'

    for data in raw_data_list:
        # 处理"加工说明"标题条目
        if "标题" in data:
            if "加工对象" not in final_dict:
                title_text = data["标题"]["文本"]
                colon_idx = title_text.find(":")
                if colon_idx != -1:
                    final_dict["加工对象"] = title_text[colon_idx + 1:]
            continue

        text_content = data["文本"]

        # 匹配尺寸信息
        if re.search(size_pattern, text_content):
            final_dict["尺寸"] = text_content
            continue

        # 匹配重量信息
        colon_idx = text_content.find(":")
        if colon_idx != -1:
            key = text_content[:colon_idx]
            value = text_content[colon_idx + 1:]
            if "GW" in key:
                final_dict["重量"] = value
            elif key not in final_dict:
                # 孔加工注释
                final_dict[key] = value
            continue

        # 排除纯数字文本
        try:
            float(text_content)
            continue
        except:
            # 未分类注释
            if text_content and text_content not in final_dict["其它未注解按3D涂色加工"]:
                final_dict["其它未注解按3D涂色加工"].append(text_content)

    # 对未分类注释去重
    final_dict["其它未注解按3D涂色加工"] = list(set(final_dict["其它未注解按3D涂色加工"]))
    return final_dict


def get_size(text):
    """获取尺寸"""
    try:
        pattern = r'(\d+(?:\.\d+)?)L[x*](\d+(?:\.\d+)?)W[x*](\d+(?:\.\d+)?)T'
        match = re.search(pattern, text)
        if match:
            length = float(match.group(1))
            width = float(match.group(2))
            height = float(match.group(3))
            return (length, width, height)
        else:
            return None
    except Exception as e:
        # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
        exc_info = traceback.extract_tb(e.__traceback__)[-1]
        err_msg = (f"[get_size] 异常：{str(e)} | "
                   f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
        print(err_msg)
        return None


def format_processes(crafts, crafts_map, sort_rules):
    # 筛选并映射：仅保留工艺且排序规则包含的项
    valid_items = []
    for craft_name, value in crafts.items():
        craft_code = crafts_map.get(craft_name)
        if not craft_code:
            continue  # 无对应代号则跳过

        # 直接使用具体代号匹配排序规则
        if craft_code in sort_rules:
            # 记录工艺名、值、在排序规则中的索引（用于排序）
            valid_items.append((sort_rules.index(craft_code), craft_name, value))

    # 按排序规则的顺序排序
    valid_items.sort(key=lambda x: x[0])

    # 拼接成指定格式：key(value)->key(value)->...
    result = "->".join([f"{name}{val}" for _, name, val in valid_items])
    result += "->质检"

    return result


def process_nx_crafts(workPart, judgement_M=False, ):
    """
    Args:
        workPart: NX当前工作部件（必填）
        judgement_M: 是否通过零件名称判断研磨需求（可选，默认False）
    Returns:
        格式化后的工艺字符串（异常时返回空字符串）
    """
    # 定义工艺-符号映射（固定配置）
    crafts_map = {
        "热处理": "PTR", "快走丝": "W/C", "中走丝": "W/Z", "慢走丝": "W/E", "钻孔": "Z", "放电加工": "EDM",
        "小磨床": "YM", "大水磨": "M", "CNC开粗": "S",
        "CNC精铣": "SS", "质检": "QC"
    }
    # 步骤1.提取加工说明信息
    text_dict = extract_and_process_notes(workPart)
    try:
        need_template = {"材质": None, "零件名称": None, "编号": None, "长/mm": None, "宽/mm": None, "厚/mm": None,
                         "数量": None, "重量/kg": None, "热处理": None, "工艺": None}
        # 步骤2.工艺提取并排序
        # 1. 获取图纸纯数字尺寸注释
        # 处理尺寸
        size = get_size(text_dict.get("尺寸", ""))
        if size:
            need_template["长/mm"] = size[0]
            need_template["宽/mm"] = size[1]
            need_template["厚/mm"] = size[2]
            # 计算重量
            need_template["重量/kg"] = f"{size[0] * size[1] * size[2] * 0.0000785:.2f}"

        num_text = get_num_notes(workPart)
        if not num_text:
            num_text = size
        # 2. 提取加工工艺和工艺属性
        crafts, _is = get_craft(workPart, text_dict, num_text, judgement_M)
        # 3. 获取工艺排序规则
        sort_rule = get_sort_rule(_is)
        # 4. 格式化工艺字符串
        formatted_crafts = format_processes(crafts, crafts_map, sort_rule)
        # 步骤5.提取零件规格参数
        need_template["工艺"] = formatted_crafts
        need_template["材质"] = get_material(text_dict.get("尺寸", ""))

        # 处理加工对象
        if "加工对象" in text_dict:
            object_split = text_dict["加工对象"].split("_")
            if len(object_split) >= 2:
                need_template["零件名称"] = object_split[0][1:-1] if len(object_split[0]) > 2 else object_split[0]
                need_template["编号"] = object_split[1]

        # 处理数量
        if "尺寸" in text_dict:
            idx = text_dict["尺寸"].find("PCS")
            if idx != -1:
                quantity_str = text_dict["尺寸"][:idx].split()[-1]
                try:
                    int(quantity_str)
                    need_template["数量"] = quantity_str
                except:
                    need_template["数量"] = None

        # 处理热处理
        HRC = []
        if "尺寸" in text_dict:
            idx = text_dict["尺寸"].find("HRC")
            if idx != -1:
                hrc_part = text_dict["尺寸"][idx:].split()[0]
                if hrc_part:
                    HRC.append(hrc_part)

        if "其它未注解按3D涂色加工" in text_dict:
            for data in text_dict["其它未注解按3D涂色加工"]:
                if "HRC" in data:
                    hrc_part = data.split()[0]
                    if hrc_part and hrc_part not in HRC:
                        HRC.append(hrc_part)

        if HRC:
            need_template["热处理"] = ",".join(HRC) if len(HRC) > 1 else HRC[0]

        return need_template, _is, sort_rule, crafts_map

    except Exception as e:
        # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
        exc_info = traceback.extract_tb(e.__traceback__)[-1]
        err_msg = (f"[process_nx_crafts] 异常：{str(e)} | "
                   f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
        print(err_msg)
        return {}


def open_part(session, part_path):
    """保证返回 NXOpen.Part 类型"""
    try:
        result = session.Parts.Open(part_path)

        # tuple unpack
        if isinstance(result, tuple):
            part = result[0]
        else:
            part = result

        if isinstance(part, NXOpen.Part):
            return part

        raise TypeError("Open() 未返回 NXOpen.Part 对象")

    except Exception as e:
        # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
        exc_info = traceback.extract_tb(e.__traceback__)[-1]
        err_msg = (f"[open_part] 异常：{str(e)} | "
                   f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
        print(err_msg)
        raise  # 重新抛出异常，让调用者处理


def dict_to_excel(data_dict, save_path):
    """
    将嵌套字典数据存入Excel文件（固定列顺序），并设置自适应列宽
    :param data_dict: 输入数据（允许字段为None，代表未提取到）
    :param save_path: 保存路径（精确到.xlsx，如"output/零件数据.xlsx"）
    """
    try:
        crafts_map = {
            "热处理": "PTR", "快走丝": "W/C", "中走丝": "W/Z", "慢走丝": "W/E",
            "钻孔": "Z", "放电加工": "EDM", "小磨床": "YM", "大水磨": "M",
            "CNC开粗": "S", "CNC精铣": "SS", "质检": "QC"
        }
        # 提取crafts_map的键值对列表（固定顺序）
        crafts_list = list(crafts_map.items())  # [(工艺名1, 代号1), (工艺名2, 代号2), ...]

        # 前置检查：确保保存目录存在
        save_dir = os.path.dirname(save_path)
        if save_dir and not os.path.exists(save_dir):
            os.makedirs(save_dir, exist_ok=True)
            print(f"保存目录不存在，已自动创建：{save_dir}")

        # 1. 创建工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "零件参数"

        # 2. 固定列顺序（保留“工艺名”“代号”列用于展示crafts_map）
        column_names = [
            "序号", "文件名称", "零件名称", "编号", "材质", "长/mm", "宽/mm", "厚/mm", "数量",
            "重量/kg", "热处理", "开粗量", "排序规则", "工艺",  "初排工艺", "工艺名", "代号",
        ]

        # 3. 写入表头（第一行）
        for col_idx, col_name in enumerate(column_names, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # 4. 写入数据（从第二行开始，按序号升序排列）
        row_idx = 2
        sorted_items = sorted(data_dict.items(), key=lambda x: int(x[0]))
        for seq_num, inner_dict in sorted_items:
            ws.cell(row=row_idx, column=1, value=seq_num)

            # 遍历除“序号”外的其他列
            for col_idx, col_name in enumerate(column_names[1:], start=2):
                # 处理普通列：None显示为空字符串
                if col_name not in ["工艺名", "代号"]:
                    value = inner_dict.get(col_name, "")
                    if value is None:
                        value = ""

                # 处理“工艺名”列：固定展示crafts_map的键（按顺序填充）
                elif col_name == "工艺名":
                    crafts_idx = row_idx - 2  # 数据从第2行开始，对应crafts_list的索引从0开始
                    value = crafts_list[crafts_idx][0] if crafts_idx < len(crafts_list) else ""

                # 处理“代号”列：固定展示crafts_map的值（按顺序填充）
                elif col_name == "代号":
                    crafts_idx = row_idx - 2
                    value = crafts_list[crafts_idx][1] if crafts_idx < len(crafts_list) else ""

                ws.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

        # 5. 设置自适应列宽（自动适配所有列）
        for col_idx in range(1, len(column_names) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[col_letter]:
                cell_value = str(cell.value) if cell.value is not None else ""
                # 中文字符宽度特殊处理（gbk编码更贴近Excel显示宽度）
                cell_length = len(cell_value.encode('gbk'))
                if cell_length > max_length:
                    max_length = cell_length
            # 列宽留余量，最小宽度8
            adjusted_width = max(max_length + 2, 8)
            ws.column_dimensions[col_letter].width = adjusted_width

        # 6. 保存文件
        wb.save(save_path)
        print(f"Excel文件已成功保存至：{save_path}（工艺名/代号列已固定展示crafts_map内容）")
    except Exception as e:
        exc_info = traceback.extract_tb(e.__traceback__)[-1]
        err_msg = (f"[dict_to_excel] 异常：{str(e)} | "
                   f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
        print(err_msg)


def process_part(
        is_heat_treatment=True,  # 是否需要热处理，默认True
        processing_surface="一面加工",  # 加工面数，可选"一面加工"或"两面加工"
        need_finish_milling=True,  # 是否需要精铣，默认True
        need_wire_cutting=True,  # 是否需要线割，默认True
        rough_deformation_large=True,  # 开粗变形量大，默认True
        is_roughing_same_surface_as_drilling=True,  # 开粗可与钻孔同面
        is_profile_surface_same_as_drilling=True,  # 型面可与钻孔同面
        wire_cutting_type=0) -> list:
    """
    基于决策树生成零件加工工艺流程

    Args:
        is_heat_treatment: 是否需要热处理
        processing_surface: 加工面数（"一面加工"或"两面加工"）
        need_finish_milling: 是否需要精铣
        need_wire_cutting: 是否需要线割
        rough_deformation_large: 开粗变形量是否大
        is_roughing_same_surface_as_drilling: 开粗与钻孔是否同面
        is_profile_surface_same_as_drilling: 型面与钻孔是否同面

    Returns:
        str: 完整工艺流程描述
    """
    sort_rule = None
    # 需要热处理
    if is_heat_treatment:
        # 一面和二面逻辑-结果都相同
        # 需要精铣
        if need_finish_milling:
            # 需要线割
            if need_wire_cutting:
                sort_rule = ['CNC开粗(0.8/0.5mm)', '热处理(0.0mm)', '大水磨(0.0mm)', '中走丝(0.0mm)', '精铣(0.0mm)',
                             '质检']
            else:
                sort_rule = ['CNC开粗(0.8/0.5mm)', '热处理(0.0mm)', '大水磨(0.0mm)', '精铣(0.0mm)', '质检']
        else:
            # 需要线割
            if need_wire_cutting:
                sort_rule = ['CNC开粗(0.8/0.5mm)', '热处理(0.0mm)', '大水磨(0.0mm)', '中走丝(0.0mm)', '质检']
            else:
                sort_rule = ['CNC开粗(0.8/0.5mm)', '热处理(0.0mm)', '大水磨(0.0mm)', '质检']

    else:
        if processing_surface == "一面加工":
            if need_finish_milling:
                # 需要线割
                if need_wire_cutting:
                    # 开粗变形量大
                    if rough_deformation_large:
                        if wire_cutting_type == 0:
                            sort_rule = ['CNC开粗(侧面0.8mm 正面0.5mm)', '大水磨(0.0mm)', '快走丝(0.0mm)',
                                         'CNC精铣(0.0mm)', '质检']
                        elif wire_cutting_type == 1:
                            sort_rule = ['CNC开粗(侧面0.8mm 正面0.5mm)', '大水磨(0.0mm)', '中走丝(0.0mm)',
                                         'CNC精铣(0.0mm)', '质检']
                        else:
                            sort_rule = ['CNC开粗(侧面0.8mm 正面0.5mm)', '大水磨(0.0mm)', '慢走丝(0.0mm)',
                                         'CNC精铣(0.0mm)', '质检']
                    else:
                        sort_rule = ['钻孔(0.0mm)', '开粗(0.0mm)', '大水磨(0.0mm)', '中走丝(0.0mm)', 'CNC精铣(0.0mm)',
                                     '质检']
                else:
                    # 开粗形变量大
                    if rough_deformation_large:
                        sort_rule = ['CNC开粗(侧面0.8mm 正面0.5mm)', '大水磨(0.0mm)', 'CNC精铣(0.0mm)', '质检']
                    else:
                        sort_rule = ['钻孔(0.0mm)', '开粗(侧面0.0mm 正面0.0mm)', '大水磨(0.0mm)', 'CNC精铣(0.0mm)', '质检']

            else:
                # 不需要线割
                if need_wire_cutting:
                    sort_rule = ['CNC开粗(侧面0.8mm 正面0.5mm)', '大水磨(0.0mm)', '快走丝(0.0mm)', '质检']
                else:
                    sort_rule = ['钻孔(0.0mm)', '开粗(侧面0.0mm 正面0.0mm)', '大水磨(0.0mm)', '质检']

        else:
            # 不需要精铣
            if not need_finish_milling:
                return ['开粗(0.0mm)', '大水磨(0.0mm)', '质检']
            # 需要线割
            if need_wire_cutting:
                # 开粗变形量大
                if rough_deformation_large:
                    # 开粗可与钻孔同面
                    if is_roughing_same_surface_as_drilling:
                        sort_rule = ["只先加工钻铣一面"]
                    else:
                        sort_rule = ['CNC开粗(侧面0.8mm 正面0.5mm)', '大水磨(0.0mm)', '中走丝(0.0mm)',
                                     'CNC精铣(0.0mm)', '质检']
                else:
                    # 型面可与钻孔同面
                    if is_profile_surface_same_as_drilling:
                        sort_rule = ['钻孔(0.0mm)', '开粗(侧面0.0mm 正面0.0mm)', '大水磨(0.0mm)', '快走丝(0.0mm)', '质检']
                    else:
                        sort_rule = ['钻孔(0.0mm)', '开粗(侧面0.0mm 正面0.0mm)', '大水磨(0.0mm)', '快走丝(0.0mm)',
                                     'CNC精铣(0.0mm)', '质检']
            else:
                # 开粗形变量大
                if rough_deformation_large:
                    sort_rule = ['CNC开粗(侧面0.8mm 正面0.5mm)', '大水磨(0.0mm)', 'CNC精铣(0.0mm)', '质检']
                else:
                    sort_rule = ['大水磨(0.0mm)', 'CNC精铣(0.0mm)', '质检']

    return sort_rule


def reorder(tech_str: str, sort_rule: list) -> str:
    """
    根据排序规则处理工艺字符串：
    - 匹配工艺替换为规则内容+按规则顺序排列
    - 慢/中/快走丝统一替换为规则中的走丝项（避免重复）
    - 不匹配工艺保留原始相对位置
    """

    def _extract_name(process: str) -> str:
        """提取工艺项的名称（去掉括号及内容）"""
        match = re.match(r'^([^(]+)', process.strip())
        return match.group(1).strip() if match else process.strip()

    def _is_wire_cutting(process_name: str) -> bool:
        """判断是否是走丝工艺（快/中/慢走丝）"""
        return any(kw in process_name for kw in ['快走丝', '中走丝', '慢走丝'])

    # -------------------------- 步骤1：解析排序规则 --------------------------
    # 拆分规则为项列表，记录“规则名称-规则完整内容”映射
    rule_items = [item.strip() for item in sort_rule]
    rule_info = [(_extract_name(item), item) for item in rule_items]  # (名称, 完整内容)

    # 识别规则中的走丝项（仅取第一个）
    rule_wire = None  # 格式：(规则走丝名称, 规则走丝完整内容)
    for name, item in rule_info:
        if _is_wire_cutting(name):
            rule_wire = (name, item)
            break

    # 构建普通工艺的“规则名称-规则内容”映射（排除走丝项）
    common_rule_map = {name: item for name, item in rule_info
                       if not rule_wire or name != rule_wire[0]}

    # -------------------------- 步骤2：替换原始工艺项 --------------------------
    orig_items = [item.strip() for item in tech_str.split('->')]
    replaced_items = []  # 替换后的工艺项列表
    wire_added = False  # 标记走丝项是否已添加（避免重复）

    for item in orig_items:
        orig_name = _extract_name(item)
        is_wire = _is_wire_cutting(orig_name)

        # 处理走丝项：统一替换为规则中的走丝项，仅添加一次
        if is_wire:
            if rule_wire and not wire_added:
                replaced_items.append(rule_wire[1])
                wire_added = True
            continue  # 跳过重复的走丝项

        # 处理普通工艺：匹配规则则替换，否则保留原内容
        matched_rule = None
        for rule_name, rule_item in common_rule_map.items():
            # 宽松匹配：规则名称与原始名称互相包含即视为匹配
            if rule_name in orig_name or orig_name in rule_name:
                matched_rule = rule_item
                break
        replaced_items.append(matched_rule if matched_rule else item)

    # -------------------------- 步骤3：按规则排序+整合非匹配项 --------------------------
    # 1. 分离“匹配规则的工艺”和“不匹配的工艺”
    rule_item_set = set(rule_items)
    matched_in_replaced = [item for item in replaced_items if item in rule_item_set]
    non_matched = [item for item in replaced_items if item not in rule_item_set]

    # 2. 按规则顺序生成匹配项的排序结果
    sorted_matched = []
    for rule_item in rule_items:
        if rule_item in matched_in_replaced:
            sorted_matched.append(rule_item)
            matched_in_replaced.remove(rule_item)  # 避免重复添加

    # 3. 将不匹配的工艺插回原始相对位置
    result = sorted_matched.copy()
    for non_item in non_matched:
        # 找到非匹配项在原始替换列表中的位置
        idx = replaced_items.index(non_item)
        # 找到前后最近的匹配项，确定插入位置
        prev_match = next((replaced_items[i] for i in range(idx - 1, -1, -1)
                           if replaced_items[i] in rule_item_set), None)
        next_match = next((replaced_items[i] for i in range(idx + 1, len(replaced_items))
                           if replaced_items[i] in rule_item_set), None)

        if prev_match and next_match:
            insert_pos = result.index(prev_match) + 1
        elif prev_match:
            insert_pos = result.index(prev_match) + 1
        elif next_match:
            insert_pos = result.index(next_match)
        else:
            insert_pos = 0  # 无匹配项时插在开头
        result.insert(insert_pos, non_item)

    return '->'.join(result)


def prt_to_dict(idx, theSession, workPart, results, label):
    """
    处理prt文件-程序入口
    """

    try:
        if not workPart:
            return None
        theSession.Parts.SetDisplay(workPart, False, False)
        theSession.Parts.SetWork(workPart)
        result, _is, sort_rule1, crafts_map = process_nx_crafts(workPart, judgement_M=False)
        sort_rule2 = process_part(
            is_heat_treatment=_is["is_PTR"],  # 是否需要热处理，默认True
            processing_surface=_is["sided_num"],  # 加工面数，可选"一面加工"或"两面加工"
            need_finish_milling=_is["is_SS"],  # 是否需要精铣，默认True
            need_wire_cutting=_is["is_WE"],  # 是否需要线割，默认True
            rough_deformation_large=True if label == 'class_B' else False,  # 开粗变形量大，默认True
            is_roughing_same_surface_as_drilling=_is["开粗与钻孔同面"],  # 开粗可与钻孔同面
            is_profile_surface_same_as_drilling=_is["型面与钻孔同面"],  # 型面可与钻孔同面
            wire_cutting_type=_is["WE_TYPE"]
        )
        result["文件名称"] = workPart.Name
        result["初排工艺"] = result['工艺']
        result["开粗量"] = label
        if sort_rule2 is None:
            result["排序规则"] = None
            results[idx] = result
            return results
        if sort_rule2[0] == "只先加工钻铣一面":
            result["排序规则"] = "未提取到具体排序规则，只先加工钻铣一面"
            results[idx] = result
            return results
        result["工艺"] = reorder(result['工艺'], sort_rule2)
        result["排序规则"] = "->".join(sort_rule2)
        results[idx] = result
        return results

    except Exception as e:
        # 统一异常格式：异常信息 + 位置（文件名、行号、函数名）
        exc_info = traceback.extract_tb(e.__traceback__)[-1]
        err_msg = (f"[batch_prt_to_excel] 异常：{str(e)} | "
                   f"异常位置：文件{exc_info.filename}，行{exc_info.lineno}，函数{exc_info.name}")
        print(err_msg)
