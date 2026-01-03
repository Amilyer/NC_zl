import os
import traceback
from datetime import datetime
from openpyxl import Workbook, load_workbook


class ExceptionLogger:
    def __init__(self, excel_path="exception_log.xlsx"):
        self.excel_path = excel_path
        self._init_excel()

    def _init_excel(self):
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "ExceptionLog"
            ws.append([
                "Time",
                "ExceptionType",
                "ExceptionMessage",
                "File",
                "Line",
                "Function",
                "Traceback"
            ])
            wb.save(self.excel_path)

    def log_exception(self, exc: Exception):
        tb = traceback.extract_tb(exc.__traceback__)
        last_tb = tb[-1] if tb else None

        time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        exc_type = type(exc).__name__
        exc_msg = str(exc)
        file_name = last_tb.filename if last_tb else ""
        line_no = last_tb.lineno if last_tb else ""
        func_name = last_tb.name if last_tb else ""
        tb_str = traceback.format_exc()

        wb = load_workbook(self.excel_path)
        ws = wb.active
        ws.append([
            time_str,
            exc_type,
            exc_msg,
            file_name,
            line_no,
            func_name,
            tb_str
        ])
        wb.save(self.excel_path)